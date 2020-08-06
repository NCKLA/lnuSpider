# -*- coding: utf-8 -*-
import random
import re
from scrapy import item
from scrapy.http import Request
from openpyxl import load_workbook
import time
import scrapy
from jqka_capital_operation.items import JqkaCapitalOperationItem


class JqkaCapitalOperationSpiderSpider(scrapy.Spider):
    name = 'jqka_capital_operation_spider'
    allowed_domains = ['basic.10.jqka.com']
    start_urls = ['http://basic.10jqka.com.cn/603221/capital.html']

    def parse(self, response):
        book = load_workbook(filename=r"C:\python\lnuSpider\data\exel\com_list.xlsx")
        sheet = book.active
        data = []
        data1 = []
        data2 = []
        data3 = []
        row_num = 1
        while row_num <= 3815:
            # 将表中第一列的1-100行数据写入data数组中
            data.append(sheet.cell(row=row_num, column=3).value)
            data1.append(sheet.cell(row=row_num, column=1).value)
            data3.append(sheet.cell(row=row_num, column=2).value)
            data2.append(row_num)
            row_num = row_num + 1
        for i in data2:
            a = str(data[i - 1])
            listedCompany_url = 'http://basic.10jqka.com.cn/' + a + '/capital.html'
            company_co = JqkaCapitalOperationItem()
            company_co['listedCompany_url'] = listedCompany_url
            listedCompany_id = data1[i - 1]
            company_co['listedCompany_id'] = listedCompany_id
            listedCompany_name = data3[i - 1]
            company_co['listedCompany_name'] = listedCompany_name
            yield scrapy.Request(company_co['listedCompany_url'], meta={'company_co': company_co}, callback=self.detail_ni, dont_filter=True)

    def detail_ni(self, response):
        # 获得相应公司对应的公司信息
        company_co = response.meta['company_co']
        item = JqkaCapitalOperationItem()

        # 公司信息
        item['listedCompany_id'] = company_co['listedCompany_id']
        item['listedCompany_name'] = company_co['listedCompany_name']
        item['listedCompany_url'] = company_co['listedCompany_url']

        # 1.募集资金来源 模块
        # listedCompany_capitalOperation_fundRaisingSource
        item['listedCompany_capitalOperation_fundRaisingSource'] = list()
        raise_trs = response.xpath("//div[@id='raise']//tbody/tr")
        for raise_tr in raise_trs:
            # 通过字典来封装对象
            raise_dict = dict()
            raise_dict['listedCompany_capitalOperation_fundRaisingSource_announcementDate'] = raise_tr.xpath("./td[1]/text()").get().strip()
            raise_dict['listedCompany_capitalOperation_fundRaisingSource_issueCategory'] = raise_tr.xpath("./td[2]/text()").get().strip()
            raise_dict['listedCompany_capitalOperation_fundRaisingSource_issueStartDate'] = raise_tr.xpath("./td[3]/text()").get().strip()
            raise_dict['listedCompany_capitalOperation_fundRaisingSource_actuallyRaisedFundsNetAmount'] = raise_tr.xpath("./td[4]/text()").get().strip()
            raise_dict['listedCompany_capitalOperation_fundRaisingSource_remainingRaisedFundsDeadline'] = raise_tr.xpath("./td[5]/text()").get().strip()
            raise_dict['listedCompany_capitalOperation_fundRaisingSource_remainingRaisedCapital'] = raise_tr.xpath("./td[6]/text()").get().strip()
            raise_dict['listedCompany_capitalOperation_fundRaisingSource_raisedFundsUtilizationRate'] = raise_tr.xpath("./td[7]/text()").get().strip()
            # 将封装好的对象添加到该模块的列表中
            item['listedCompany_capitalOperation_fundRaisingSource'].append(raise_dict)


        # 2.项目投资 模块
        # listedCompany_capitalOperation_projectInvestment
        item['listedCompany_capitalOperation_projectInvestment'] = list()
        progress_trs = response.xpath("//div[@id='progress']//tbody/tr")
        for progress_tr in progress_trs:
            # 通过字典来封装对象
            progress_dict = dict()
            progress_dict['listedCompany_capitalOperation_projectInvestment_announcementDate'] = progress_tr.xpath("./td[1]/text()").get().strip()
            progress_dict['listedCompany_capitalOperation_projectInvestment_projectName'] = progress_tr.xpath("./td[2]/text()").get().strip()
            progress_dict['listedCompany_capitalOperation_projectInvestment_commitmentUseRaisedFunds'] = progress_tr.xpath("./td[3]/text()").get().strip()
            progress_dict['listedCompany_capitalOperation_projectInvestment_raisedFundsInvested'] = progress_tr.xpath("./td[4]/text()").get().strip()
            progress_dict['listedCompany_capitalOperation_projectInvestment_constructionPeriod'] = progress_tr.xpath("./td[5]/text()").get().strip()
            progress_dict['listedCompany_capitalOperation_projectInvestment_afterTaxYield'] = progress_tr.xpath("./td[6]/text()").get().strip()
            progress_dict['listedCompany_capitalOperation_projectInvestment_forecastAnnualNewNetProfit'] = progress_tr.xpath("./td[7]/text()").get().strip()
            if progress_tr.xpath("./td[8]/a/@content").get():
                progress_dict['listedCompany_capitalOperation_projectInvestment_projectBriefIntroduction'] = progress_tr.xpath("./td[8]/a/@content").get().strip()
            else:
                progress_dict['listedCompany_capitalOperation_projectInvestment_projectBriefIntroduction'] = progress_tr.xpath("./td[8]/text()").get().strip()
                # 将封装好的对象添加到该模块的列表中
            item['listedCompany_capitalOperation_projectInvestment'].append(progress_dict)


        # 3.收购兼并 模块
        # listedCompany_capitalOperation_purchaseandMergers
        item['listedCompany_capitalOperation_purchaseandMergers'] = list()
        asset_tables = response.xpath("//div[@id='assetdata']/table")
        for asset_table in asset_tables:
            # 通过字典来封装对象
            asset_dict = dict()
            asset_dict['listedCompany_capitalOperation_purchaseandMergers_announcementDate'] = asset_table.xpath(".//tr[1]/td[1]/em/text()").get().strip()
            asset_dict['listedCompany_capitalOperation_purchaseandMergers_transactionAmount'] = asset_table.xpath(".//tr[1]/td[2]/text()").get().strip()
            asset_dict['listedCompany_capitalOperation_purchaseandMergers_transactionSchedule'] = asset_table.xpath(".//tr[1]/td[3]/text()").get().strip()
            asset_dict['listedCompany_capitalOperation_purchaseandMergers_transactionTarget'] = asset_table.xpath(".//tr[2]//p/text()").get().strip()
            asset_dict['listedCompany_capitalOperation_purchaseandMergers_buyer'] = asset_table.xpath(".//tr[3]/td/text()").get().strip()
            asset_dict['listedCompany_capitalOperation_purchaseandMergers_seller'] = asset_table.xpath(".//tr[4]/td/text()").get().strip()
            asset_dict['listedCompany_capitalOperation_purchaseandMergers_transactionOverview'] = asset_table.xpath(".//tr[5]//p/text()").get().strip()
            # 将封装好的对象添加到该模块的列表中
            item['listedCompany_capitalOperation_purchaseandMergers'].append(asset_dict)


        # 4.股权投资 模块
        # listedCompany_capitalOperation_equityInvestment
        item['listedCompany_capitalOperation_equityInvestment'] = list()
        # 日期数等于日期对应的div数目 len(invest_dates) = len(invest_divs)
        invest_dates = response.xpath("//div[@id='invest']//a[@class='fdates']/text()").getall()
        invest_divs = response.xpath("//div[@id='invest']//div[@class='m_tab_content']")
        for i in range(0, len(invest_divs)):
            # 每个invest_div(每个日期)里面有N个invest_tr(1个invest_trs列表)
            invest_trs = invest_divs[i].xpath("./table[1]/tbody/tr")
            for invest_tr in invest_trs:
                # 通过字典来封装对象
                invest_dict = dict()
                # 日期 与 其他字段有所不同，直接从 invest_dates 列表 根据下标取就行
                invest_dict['listedCompany_capitalOperation_equityInvestment_date'] = invest_dates[i]
                invest_dict['listedCompany_capitalOperation_equityInvestment_investmentType'] = invest_tr.xpath("./td[1]/text()").get().strip()
                invest_dict['listedCompany_capitalOperation_equityInvestment_holdingsNumber'] = invest_tr.xpath("./td[2]/text()").get().strip()
                invest_dict['listedCompany_capitalOperation_equityInvestment_accumulatedInitialInvestment'] = invest_tr.xpath("./td[3]/text()").get().strip()
                invest_dict['listedCompany_capitalOperation_equityInvestment_cumulativeEndingFaceValue'] = invest_tr.xpath("./td[4]/text()").get().strip()
                invest_dict['listedCompany_capitalOperation_equityInvestment_quarterEndedProfitAndLoss'] = invest_tr.xpath("./td[5]/span/text()").get().strip()
                invest_dict['listedCompany_capitalOperation_equityInvestment_performanceImpact'] = invest_tr.xpath("./td[6]/text()").get().strip()
                # 将封装好的对象添加到该模块的列表中
                item['listedCompany_capitalOperation_equityInvestment'].append(invest_dict)


        # 5.股权转让 模块
        # listedCompany_capitalOperation_equityTransfer
        item['listedCompany_capitalOperation_equityTransfer'] = list()
        transfer_tables = response.xpath("//div[@id='transferdata']/table")
        for transfer_table in transfer_tables:
            # 通过字典来封装对象
            transfer_dict = dict()
            transfer_dict['listedCompany_capitalOperation_equityTransfer_announcementDate'] = transfer_table.xpath(".//tr[1]//td[1]/em/text()").get().strip()
            transfer_dict['listedCompany_capitalOperation_equityTransfer_transactionAmount'] = transfer_table.xpath(".//tr[1]//td[2]/text()").get().strip()
            transfer_dict['listedCompany_capitalOperation_equityTransfer_transferRatio'] = transfer_table.xpath(".//tr[1]//td[3]/text()").get().strip()
            transfer_dict['listedCompany_capitalOperation_equityTransfer_transferor'] = transfer_table.xpath(".//tr[2]/td[1]/text()").get().strip()
            transfer_dict['listedCompany_capitalOperation_equityTransfer_transactionTarget'] = transfer_table.xpath(".//tr[2]/td[2]/text()").get().strip()
            transfer_dict['listedCompany_capitalOperation_equityTransfer_assignee'] = transfer_table.xpath(".//tr[3]/td[1]/text()").get().strip()
            if transfer_table.xpath(".//tr[3]/td[2]//tbody//td"):
                transfer_dict['listedCompany_capitalOperation_equityTransfer_transactionBriefIntroduction'] = transfer_table.xpath(".//tr[3]/td[2]//tbody//td/text()").get().strip()
            else:
                transfer_dict['listedCompany_capitalOperation_equityTransfer_transactionBriefIntroduction'] = "-"
            transfer_dict['listedCompany_capitalOperation_equityTransfer_transactionImpact'] = transfer_table.xpath(".//tr[4]/td/text()").get().strip()
            # 将封装好的对象添加到该模块的列表中
            item['listedCompany_capitalOperation_equityTransfer'].append(transfer_dict)


        # 6.关联交易 模块
        # listedCompany_capitalOperation_relatedTransactions
        item['listedCompany_capitalOperation_relatedTransactions'] = list()
        connect_tables = response.xpath("//div[@id='connectdata']/table")
        for connect_table in connect_tables:
            # 通过字典来封装对象
            connect_dict = dict()
            connect_dict['listedCompany_capitalOperation_relatedTransactions_announcementDate'] = connect_table.xpath(".//tr[1]/td[1]/em/text()").get().strip()
            connect_dict['listedCompany_capitalOperation_relatedTransactions_transactionAmount'] = connect_table.xpath(".//tr[1]/td[2]/text()").get().strip()
            connect_dict['listedCompany_capitalOperation_relatedTransactions_paymentMethod'] = connect_table.xpath(".//tr[1]/td[3]/text()").get().strip()
            connect_dict['listedCompany_capitalOperation_relatedTransactions_tradingParty'] = connect_table.xpath(".//tr[2]/td[1]/text()").get().strip()
            connect_dict['listedCompany_capitalOperation_relatedTransactions_tradingMethod'] = connect_table.xpath(".//tr[2]/td[2]/text()").get().strip()
            connect_dict['listedCompany_capitalOperation_relatedTransactions_correlation'] = connect_table.xpath(".//tr[3]/td/text()").get().strip()
            if connect_table.xpath(".//tr[4]//p/text()").get():
                connect_dict[
                    'listedCompany_capitalOperation_relatedTransactions_transactionBriefIntroduction'] = connect_table.xpath(".//tr[4]//p/text()").get().strip()
            # 将封装好的对象添加到该模块的列表中
            item['listedCompany_capitalOperation_relatedTransactions'].append(connect_dict)


        yield item

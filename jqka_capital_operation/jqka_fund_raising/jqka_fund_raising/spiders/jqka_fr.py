# -*- coding: utf-8 -*-
import random
import time

import scrapy
from openpyxl import load_workbook
from scrapy import Request

from jqka_capital_operation.jqka_fund_raising.jqka_fund_raising.items import JqkaFundRaisingItem


class JqkaFrSpider(scrapy.Spider):
    name = 'jqka_fr'
    allowed_domains = ['basic.10.jqka.com']


    def start_requests(self):
        yield Request("http://basic.10jqka.com.cn/603221/capital.html", headers={
            'User-Agent': "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.87 Safari/537.36"})

    def parse(self, response):
         book = load_workbook(filename="com_list.xlsx")
         sheet = book.active
         data = []
         data1 = []
         data2 = []
         data3 = []
         row_num = 1
         while row_num <= 500:
            #将表中第一列的1-100行数据写入data数组中
            data.append(sheet.cell(row=row_num, column=3).value)
            data1.append(sheet.cell(row=row_num, column=1).value)
            data3.append(sheet.cell(row=row_num, column=2).value)
            data2.append(row_num)
            row_num = row_num + 1
         for i in data2:
                # url = 'http://basic.10jqka.com.cn/'+data[i]+'/company.html'
                    # print(data[i-1])
              listedCompany_url = 'http://basic.10jqka.com.cn/' + data[i - 1] + '/capital.html'
              jqka_fr1= JqkaFundRaisingItem()
              jqka_fr1['listedCompany_url'] = listedCompany_url
              listedCompany_id = data1[i - 1]
              jqka_fr1['listedCompany_id'] = listedCompany_id
              listedCompany_name = data3[i - 1]
              jqka_fr1['listedCompany_name'] = listedCompany_name
              #print(listedCompany_id)
           # pandas读取表格
           # res_elements = etree.HTML(response.text)
           # table = res_elements.xpath("//table[@class='tbody']")
           # table = etree.tostring(table[0], encoding='utf-8').decode()
           # df = pd.read_html(table, encoding='utf-8', header=0)
           # results = list(df.T.to_dict().values())  # 转换成列表嵌套字典的格式
           # print(results)
              yield scrapy.Request(jqka_fr1['listedCompany_url'],
                              meta={'jqka_fr1': jqka_fr1}, callback=self.fr, dont_filter=True)

         return


    def fr(self, response):
        print("=====准备公司中的信息====")
        jqka_fr1 = response.meta['jqka_fr1']

        print("=====信息准备完毕====")

        content1 = response.xpath("//div[@class='m_box'][@id='raise']//table/tbody/tr")
        #content1 = contents.xpath("./div[@class='bd pt5 pagination']//table/tbody/tr")


        #print(response.text)
        fund_raising = list()
        for content in content1:
            single = dict()
            listedCompany_fundRaisingSource_announcementDate = content.xpath("./td[1]/text()").getall()
            single['listedCompany_fundRaisingSource_announcementDate'] = "".join(listedCompany_fundRaisingSource_announcementDate).strip()
            listedCompany_fundRaisingSource_issueCategory = content.xpath("./td[2]/text()").getall()
            single['listedCompany_fundRaisingSource_issueCategory'] = "".join(listedCompany_fundRaisingSource_issueCategory).strip()
            listedCompany_fundRaisingSource_issueStartDate = content.xpath("./td[3]/text()").getall()
            single['listedCompany_fundRaisingSource_issueStartDate'] = "".join(listedCompany_fundRaisingSource_issueStartDate).strip()
            listedCompany_fundRaisingSource_actuallyRaisedFundsNetAmount = content.xpath("./td[43]/text()").getall()
            single['listedCompany_fundRaisingSource_actuallyRaisedFundsNetAmount'] = "".join(listedCompany_fundRaisingSource_actuallyRaisedFundsNetAmount).strip()
            listedCompany_fundRaisingSource_remainingRaisedFundsDeadline = content.xpath("./td[5]/text()").getall()
            single['listedCompany_fundRaisingSource_remainingRaisedFundsDeadline'] = "".join(listedCompany_fundRaisingSource_remainingRaisedFundsDeadline).strip()
            listedCompany_fundRaisingSource_remainingRaisedCapital = content.xpath("./td[6]/text()").getall()
            single['listedCompany_fundRaisingSource_remainingRaisedCapital'] = "".join(listedCompany_fundRaisingSource_remainingRaisedCapital).strip()
            listedCompany_fundRaisingSource_raisedFundsUtilizationRate = content.xpath("./td[7]/text()").getall()
            single['listedCompany_fundRaisingSource_raisedFundsUtilizationRate'] = "".join(listedCompany_fundRaisingSource_raisedFundsUtilizationRate).strip()


            fund_raising.append(single)
        jqka_fr1['fund_raising'] = fund_raising

        time.sleep(random.randint(2, 5))
        yield jqka_fr1



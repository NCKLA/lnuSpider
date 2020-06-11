
import random
import time

import scrapy
from openpyxl import load_workbook
from scrapy import Request

from jqka_capital_operation.jqka_fund_raising.jqka_fund_raising.items import JqkaFundRaisingItem
from jqka_capital_operation.jqka_project_investment.jqka_project_investment.items import JqkaProjectInvestmentItem


class JqkaPiSpider(scrapy.Spider):
    name = 'jqka_pi'
    allowed_domains = ['basic.10.jqka.com']


    def start_requests(self):
        yield Request("http://basic.10jqka.com.cn/603221/capital.html", headers={
            'User-Agent': "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.87 Safari/537.36"})

    def parse(self, response):
         book = load_workbook(filename="com_list.xlsx")
         sheet = book.active
         data = []
         data1 = []
         data2 = []
         data3 = []
         row_num = 1
         while row_num <= 500:
            #将表中第一列的1-100行数据写入data数组中
            data.append(sheet.cell(row=row_num, column=3).value)
            data1.append(sheet.cell(row=row_num, column=1).value)
            data3.append(sheet.cell(row=row_num, column=2).value)
            data2.append(row_num)
            row_num = row_num + 1
         for i in data2:
                # url = 'http://basic.10jqka.com.cn/'+data[i]+'/company.html'
                    # print(data[i-1])
              listedCompany_url = 'http://basic.10jqka.com.cn/' + data[i - 1] + '/capital.html'
              jqka_pi1= JqkaProjectInvestmentItem()
              jqka_pi1['listedCompany_url'] = listedCompany_url
              listedCompany_id = data1[i - 1]
              jqka_pi1['listedCompany_id'] = listedCompany_id
              listedCompany_name = data3[i - 1]
              jqka_pi1['listedCompany_name'] = listedCompany_name
              #print(listedCompany_id)
           # pandas读取表格
           # res_elements = etree.HTML(response.text)
           # table = res_elements.xpath("//table[@class='tbody']")
           # table = etree.tostring(table[0], encoding='utf-8').decode()
           # df = pd.read_html(table, encoding='utf-8', header=0)
           # results = list(df.T.to_dict().values())  # 转换成列表嵌套字典的格式
           # print(results)
              yield scrapy.Request(jqka_pi1['listedCompany_url'],
                              meta={'jqka_pi1': jqka_pi1}, callback=self.pi, dont_filter=True)

         return


    def pi(self, response):
        print("=====准备公司中的信息====")
        jqka_pi1 = response.meta['jqka_pi1']

        print("=====信息准备完毕====")

        #contents = response.xpath("//div[@class='m_box'][@id='progress']")
        #content1 = contents.xpath("./div[@class='bd pt5 pagination']//table/tbody/tr")

        content1 = response.xpath("//div[@class='m_box'][@id='progress']//table/tbody/tr")

        #print(response.text)
        project_investment = list()
        for content in content1:
            single = dict()
            listedCompany_projectInvestment_announcementDate = content.xpath("./td[1]/text()").getall()
            single['listedCompany_projectInvestment_announcementDate'] = "".join(listedCompany_projectInvestment_announcementDate).strip()
            listedCompany_projectInvestment_projectName = content.xpath("./td[2]/text()").getall()
            single['listedCompany_projectInvestment_projectName'] = "".join(listedCompany_projectInvestment_projectName).strip()
            listedCompany_projectInvestment_commitmentUseRaisedFunds = content.xpath("./td[3]/text()").getall()
            single['listedCompany_projectInvestment_commitmentUseRaisedFunds'] = "".join(listedCompany_projectInvestment_commitmentUseRaisedFunds).strip()
            listedCompany_projectInvestment_raisedFundsInvested = content.xpath("./td[43]/text()").getall()
            single['listedCompany_projectInvestment_raisedFundsInvested'] = "".join(listedCompany_projectInvestment_raisedFundsInvested).strip()
            listedCompany_projectInvestment_constructionPeriod = content.xpath("./td[5]/text()").getall()
            single['listedCompany_projectInvestment_constructionPeriod'] = "".join(listedCompany_projectInvestment_constructionPeriod).strip()
            listedCompany_projectInvestment_afterTaxYield = content.xpath("./td[6]/text()").getall()
            single['listedCompany_projectInvestment_afterTaxYield'] = "".join(listedCompany_projectInvestment_afterTaxYield).strip()
            listedCompany_projectInvestment_forecastAnnualNewNetProfit = content.xpath("./td[7]/text()").getall()
            single['listedCompany_projectInvestment_forecastAnnualNewNetProfit'] = "".join(listedCompany_projectInvestment_forecastAnnualNewNetProfit).strip()
            listedCompany_projectInvestment_projectBriefIntroduction = content.xpath("./td[8][@class='tc']/a['content']").getall()
            single['listedCompany_projectInvestment_projectBriefIntroduction'] = "".join(listedCompany_projectInvestment_projectBriefIntroduction).strip()

            project_investment.append(single)

        jqka_pi1['project_investment'] = project_investment

        time.sleep(random.randint(2, 5))
        yield jqka_pi1



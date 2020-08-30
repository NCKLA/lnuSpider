# -*- coding: utf-8 -*-
import requests
import scrapy
from jqka_Business_analysis.items import JqkaBusinessAnalysisItem
from lxml import etree
from openpyxl import load_workbook
from scrapy import item
from scrapy.http import Request
# 获取excel需要引入的包
from openpyxl import load_workbook
from selenium import webdriver
import time
import scrapy
import pandas as pd


class JqkaBaSpider(scrapy.Spider):
    name = 'jqka_Ba'
    allowed_domains = ['http://basic.10jqka.com.cn/603221/operate.html']
    # start_urls = ['http://basic.10jqka.com.cn/603221/operate.html/']

    def __init__(self):
        options = webdriver.FirefoxOptions()
        options.add_argument("--headless")  # 设置火狐为headless无界面模式
        options.add_argument("--disable-gpu")
        self.driver_path = r"D:\project\geckodriver.exe"
        self.driver = webdriver.Firefox(executable_path=self.driver_path, firefox_options=options)

    def start_requests(self):
        yield Request("http://basic.10jqka.com.cn/603221/operate.html", headers={
            'User-Agent': "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.87 Safari/537.36"})

    def parse(self, response):
        book = load_workbook(filename=r"C:\python\lnuSpider\data\exel\com_list.xlsx")
        sheet = book.active
        data = []
        data1 = []
        data2 = []
        data3 = []
        data4 = []
        row_num = 1
        while row_num <= 1:
            # 将表中第一列的1-100行数据写入data数组中
            data.append(sheet.cell(row=row_num, column=3).value)
            data1.append(sheet.cell(row=row_num, column=1).value)
            data3.append(sheet.cell(row=row_num, column=2).value)
            data4.append(sheet.cell(row=row_num, column=4).value)
            data2.append(row_num)
            row_num = row_num + 1
        for i in data2:
            # url = 'http://basic.10jqka.com.cn/'+data[i]+'/company.html'
            # print(data[i-1])
            listedCompany_url = 'https://basic.10jqka.com.cn/' + data[i - 1] + '/operate.html'
            company_ba= JqkaBusinessAnalysisItem()
            company_ba['listedCompany_url'] = listedCompany_url
            listedCompany_id = data1[i - 1]
            company_ba['listedCompany_id'] = listedCompany_id
            listedCompany_name = data3[i - 1]
            company_ba['listedCompany_name'] = listedCompany_name
            listedCompany_fullName = data4[i - 1]
            company_ba['listedCompany_fullName'] = listedCompany_fullName
            # print(listedCompany_id)
            yield scrapy.Request(company_ba['listedCompany_url'],
                                 meta={'company_ba': company_ba}, callback=self.detail_ni, dont_filter=True)
        return

    def detail_ni(self, response):
        company_ba = response.meta['company_ba']
        self.driver.get(company_ba['listedCompany_url'])
        time.sleep(10)
        print(company_ba['listedCompany_url'])
        page = self.driver.page_source
        print(page)
        # print(response.text)
        # dict1 = dict()
        # =====================主营构成分析======================
        # list1=list()
        # listedCompany_businessAnalysis_mainBusinessCompositionAnalysis=list()
        # root = response.xpath("//div[@id='analysis']/div[@class='bd pt5']/div[@class='m_tab mt15']/ul/li")
        # for x in range(len(root)):
        #     x = x + 1
        #     print(x)
        #     root2 = response.xpath("//div[@id='analysis']/div[@class='bd pt5']/div[{}]".format(x + 3))
        #     print(root2)
        #     root3 = root2.xpath("./table/tbody/tr")
        #     list0=list()
        #     listedCompany_businessAnalysis_mainBusinessCompositionAnalysis_time = response.xpath("//div[@id='analysis']/div[@class='bd pt5']/div[@class='m_tab mt15']/ul/li[{}]/a/text()".format(x)).getall()
        #     listedCompany_businessAnalysis_mainBusinessCompositionAnalysis_time = ''.join(listedCompany_businessAnalysis_mainBusinessCompositionAnalysis_time).strip()
        #     print(listedCompany_businessAnalysis_mainBusinessCompositionAnalysis_time)
        #     sin=dict()
        #     sin['listedCompany_businessAnalysis_mainBusinessCompositionAnalysis_time']=listedCompany_businessAnalysis_mainBusinessCompositionAnalysis_time
        #     list1.append(sin)
        #     for rot3 in root3:
        #         single=dict()
        #         listedCompany_businessAnalysis_mainBusinessCompositionAnalysis_bussinessName = rot3.xpath("./td[1]/text()").getall()
        #         listedCompany_businessAnalysis_mainBusinessCompositionAnalysis_bussinessName = ''.join(
        #             listedCompany_businessAnalysis_mainBusinessCompositionAnalysis_bussinessName).strip()
        #         single['listedCompany_businessAnalysis_mainBusinessCompositionAnalysis_bussinessName']=listedCompany_businessAnalysis_mainBusinessCompositionAnalysis_bussinessName
        #
        #         listedCompany_businessAnalysis_mainBusinessCompositionAnalysis_operatingIncome = rot3.xpath("./td[2]/text()").getall()
        #         listedCompany_businessAnalysis_mainBusinessCompositionAnalysis_operatingIncome = ''.join(
        #             listedCompany_businessAnalysis_mainBusinessCompositionAnalysis_operatingIncome).strip()
        #         single[
        #             'listedCompany_businessAnalysis_mainBusinessCompositionAnalysis_operatingIncome'] = listedCompany_businessAnalysis_mainBusinessCompositionAnalysis_operatingIncome
        #
        #         listedCompany_businessAnalysis_mainBusinessCompositionAnalysis_incomeRatio = rot3.xpath("./td[3]/text()").getall()
        #         listedCompany_businessAnalysis_mainBusinessCompositionAnalysis_incomeRatio = ''.join(
        #             listedCompany_businessAnalysis_mainBusinessCompositionAnalysis_incomeRatio).strip()
        #         single[
        #             'listedCompany_businessAnalysis_mainBusinessCompositionAnalysis_incomeRatio'] = listedCompany_businessAnalysis_mainBusinessCompositionAnalysis_incomeRatio
        #
        #         listedCompany_businessAnalysis_mainBusinessCompositionAnalysis_operatingCosts = rot3.xpath("./td[4]/text()").getall()
        #         listedCompany_businessAnalysis_mainBusinessCompositionAnalysis_operatingCosts = ''.join(
        #             listedCompany_businessAnalysis_mainBusinessCompositionAnalysis_operatingCosts).strip()
        #         single[
        #             'listedCompany_businessAnalysis_mainBusinessCompositionAnalysis_operatingCosts'] = listedCompany_businessAnalysis_mainBusinessCompositionAnalysis_operatingCosts
        #
        #         listedCompany_businessAnalysis_mainBusinessCompositionAnalysis_costRatio = rot3.xpath("./td[5]/text()").getall()
        #         listedCompany_businessAnalysis_mainBusinessCompositionAnalysis_costRatio = ''.join(
        #             listedCompany_businessAnalysis_mainBusinessCompositionAnalysis_costRatio).strip()
        #         single[
        #             'listedCompany_businessAnalysis_mainBusinessCompositionAnalysis_costRatio'] = listedCompany_businessAnalysis_mainBusinessCompositionAnalysis_costRatio
        #
        #         listedCompany_businessAnalysis_mainBusinessCompositionAnalysis_percentageOfProfit = rot3.xpath("./td[6]/text()").getall()
        #         listedCompany_businessAnalysis_mainBusinessCompositionAnalysis_percentageOfProfit = ''.join(
        #             listedCompany_businessAnalysis_mainBusinessCompositionAnalysis_percentageOfProfit).strip()
        #         single[
        #             'listedCompany_businessAnalysis_mainBusinessCompositionAnalysis_percentageOfProfit'] = listedCompany_businessAnalysis_mainBusinessCompositionAnalysis_percentageOfProfit
        #
        #         listedCompany_businessAnalysis_mainBusinessCompositionAnalysis_rateOfMargin = rot3.xpath("./td[7]/text()").getall()
        #         listedCompany_businessAnalysis_mainBusinessCompositionAnalysis_rateOfMargin = ''.join(
        #         listedCompany_businessAnalysis_mainBusinessCompositionAnalysis_rateOfMargin).strip()
        #         single[
        #             'listedCompany_businessAnalysis_mainBusinessCompositionAnalysis_rateOfMargin'] = listedCompany_businessAnalysis_mainBusinessCompositionAnalysis_rateOfMargin
        #         list0.append(single)
        #     list1.append(list0)
        # listedCompany_businessAnalysis_mainBusinessCompositionAnalysis.append(list1)
        # print(listedCompany_businessAnalysis_mainBusinessCompositionAnalysis)
        # company_ba['listedCompany_businessAnalysis_mainBusinessCompositionAnalysis'] = listedCompany_businessAnalysis_mainBusinessCompositionAnalysis



        # print(response.text)
        # root = response.xpath("//div[@class='m_box i_hover_box'][@id='analysis']/div[@class='bd pt5']")
        # zuixingonggaoriqi = response.xpath("//div[@class='m_box gqtz i_hover_box'][@id='invest']//div[@class='m_tab mt15 clearfix']/div/span/text()").getall()
        # print(zuixingonggaoriqi)
        # print(root)
        # time1 = root.xpath("./div[@class='m_tab mt15']/ul[1]/li[1]/a/text()").get()
        # print(time1)
        # yewumingcheng = root.xpath("./div[@class='m_tab_content']/table/tbody/tr[1]/td[1]/text()").get()
        # print(yewumingcheng)
        # y = root.xpath("./div[5]/table/tbody/tr[1]/td[2]/text()").get()
        # print(y)
        # yewumingchen = root.xpath("./div[@class='m_tab_content']/table/tbody/tr[1]/td[2]/text()").get()
        # print(yewumingchen)
        # y1 = root.xpath("./div[5]/table/tbody/tr[1]/td[3]/text()").get()
        # print(y1)
        # yewumingche = root.xpath("./div[@class='m_tab_content']/table/tbody/tr[1]/td[3]/text()").get()
        # print(yewumingche)
        # y2 = root.xpath("./div[5]/table/tbody/tr[1]/td[4]/text()").get()
        # print(y2)
        # yewumingch = root.xpath("./div[@class='m_tab_content']/table/tbody/tr[1]/td[4]/text()").get()
        # print(yewumingch)
        # y3 = root.xpath("./div[5]/table/tbody/tr[1]/td[4]/text()").get()
        # print(y3)
        # yewumingc = root.xpath("./div[@class='m_tab_content']/table/tbody/tr[1]/td[4]/text()").get()
        # print(yewumingc)
        # y4 = root.xpath("./div[5]/table/tbody/tr[1]/td[4]/text()").get()
        # print(y4)

        # ==========================主营介绍=========================
        listedCompany_businessAnalysis_mainBusiness = dict()
        root2 = response.xpath("//div[@class='content page_event_content']/div[@class='m_box main_intro i_hover_box']")
        # print(root2)
        cont1 = root2.xpath("./div[@class='bd pt5']/div[@class='mt15']/ul/li[1]/p/text()").getall()
        cont1 = ''.join(cont1).strip()
        listedCompany_businessAnalysis_mainBusiness['listedCompany_businessAnalysis_mainBusinessIntro_mainBusiness '] = cont1
        cont2 = root2.xpath("./div[@class='bd pt5']/div[@class='mt15']/ul/li[2]/p/text()").getall()
        cont2 = ''.join(cont2).strip()
        listedCompany_businessAnalysis_mainBusiness['listedCompany_businessAnalysis_mainBusinessIntro_productType '] = cont2
        cont3 = root2.xpath("./div[@class='bd pt5']/div[@class='mt15']/ul/li[3]/p/text()").getall()
        cont3 = ''.join(cont3).strip()
        listedCompany_businessAnalysis_mainBusiness['listedCompany_businessAnalysis_mainBusinessIntro_productName '] = cont3
        cont4 = root2.xpath("./div[@class='bd pt5']/div[@class='mt15']/ul/li[4]/p/text()").getall()
        listedCompany_businessAnalysis_mainBusiness['listedCompany_businessAnalysis_mainBusinessIntro_businessScope'] = cont4
        company_ba['listedCompany_businessAnalysis_mainBusiness'] = listedCompany_businessAnalysis_mainBusiness


        # root = response.xpath("//div[@class='m_box i_hover_box'][@id='analysis']/div[@class='bd pt5']/div[@class='m_tab mt15']")
        # root11 = response.xpath("//div[@class='m_box i_hover_box'][@id='analysis']/div[@class='bd pt5']/div[@class='m_tab mt15']")
        # root = response.xpath("//div[@class='m_box i_hover_box'][@id='analysis']/div[@class='bd pt5']")
        # root_trs = root.xpath("./div[@class='m_tab_content'][1]/table/tbody/tr")
        # list1 = list()
        # listedCompany_mainBusinessCompositionAnalysis_1 = dict()
        # for root_tr in root_trs:
        #     cont1 = root_tr.xpath("./td[1]/text()").getall()
        #     cont1 = ''.join(cont1).strip()
        #     listedCompany_mainBusinessCompositionAnalysis_1['listedCompany_mainBusinessCompositionAnalysis_bussinessName'] = cont1
        #     cont2 = root_tr.xpath("./td[2]/text()").getall()
        #     cont2 = ''.join(cont2).strip()
        #     listedCompany_mainBusinessCompositionAnalysis_1['listedCompany_mainBusinessCompositionAnalysis_operatingIncome'] = cont2
        #     cont3 = root_tr.xpath("./td[3]/text()").getall()
        #     cont3 = ''.join(cont3).strip()
        #     listedCompany_mainBusinessCompositionAnalysis_1['listedCompany_mainBusinessCompositionAnalysis_incomeRatio'] = cont3
        #     cont4 = root_tr.xpath("./td[4]/text()").getall()
        #     cont4 = ''.join(cont4).strip()
        #     listedCompany_mainBusinessCompositionAnalysis_1['listedCompany_mainBusinessCompositionAnalysis_operatingCosts'] = cont4
        #     cont5 = root_tr.xpath("./td[5]/text()").getall()
        #     cont5 = ''.join(cont5).strip()
        #     listedCompany_mainBusinessCompositionAnalysis_1['listedCompany_mainBusinessCompositionAnalysis_costRatio'] = cont5
        #     cont6 = root_tr.xpath("./td[6]/text()").getall()
        #     cont6 = ''.join(cont6).strip()
        #     listedCompany_mainBusinessCompositionAnalysis_1['listedCompany_mainBusinessCompositionAnalysis_percentageOfProfit'] = cont6
        #     cont7 = root_tr.xpath("./td[7]/text()").getall()
        #     cont7 = ''.join(cont7).strip()
        #     listedCompany_mainBusinessCompositionAnalysis_1['listedCompany_mainBusinessCompositionAnalysis_rateOfMargin'] = cont7
        #     listedCompany_mainBusinessCompositionAnalysis_1_time = root11.xpath("./ul/li[1]/a/text()").getall()
        #     listedCompany_mainBusinessCompositionAnalysis_1_time = ''.join(listedCompany_mainBusinessCompositionAnalysis_1_time).strip()
        #     listedCompany_mainBusinessCompositionAnalysis_1['listedCompany_mainBusinessCompositionAnalysis_time'] = listedCompany_mainBusinessCompositionAnalysis_1_time
        #     list1.append(listedCompany_mainBusinessCompositionAnalysis_1)
        # company_ba['listedCompany_mainBusinessCompositionAnalysis_1'] = list1
        #
        # root_trs1 = root.xpath("./div[@class='m_tab_content'][2]/table/tbody/tr")
        # list2 = list()
        # listedCompany_mainBusinessCompositionAnalysis_2 = dict()
        # for root_tr in root_trs1:
        #     cont1 = root_tr.xpath("./td[1]/text()").getall()
        #     cont1 = ''.join(cont1).strip()
        #     listedCompany_mainBusinessCompositionAnalysis_2[
        #         'listedCompany_mainBusinessCompositionAnalysis_bussinessName'] = cont1
        #     cont2 = root_tr.xpath("./td[2]/text()").getall()
        #     cont2 = ''.join(cont2).strip()
        #     listedCompany_mainBusinessCompositionAnalysis_2[
        #         'listedCompany_mainBusinessCompositionAnalysis_operatingIncome'] = cont2
        #     cont3 = root_tr.xpath("./td[3]/text()").getall()
        #     cont3 = ''.join(cont3).strip()
        #     listedCompany_mainBusinessCompositionAnalysis_2[
        #         'listedCompany_mainBusinessCompositionAnalysis_incomeRatio'] = cont3
        #     cont4 = root_tr.xpath("./td[4]/text()").getall()
        #     cont4 = ''.join(cont4).strip()
        #     listedCompany_mainBusinessCompositionAnalysis_2[
        #         'listedCompany_mainBusinessCompositionAnalysis_operatingCosts'] = cont4
        #     cont5 = root_tr.xpath("./td[5]/text()").getall()
        #     cont5 = ''.join(cont5).strip()
        #     listedCompany_mainBusinessCompositionAnalysis_2[
        #         'listedCompany_mainBusinessCompositionAnalysis_costRatio'] = cont5
        #     cont6 = root_tr.xpath("./td[6]/text()").getall()
        #     cont6 = ''.join(cont6).strip()
        #     listedCompany_mainBusinessCompositionAnalysis_2[
        #         'listedCompany_mainBusinessCompositionAnalysis_percentageOfProfit'] = cont6
        #     cont7 = root_tr.xpath("./td[7]/text()").getall()
        #     cont7 = ''.join(cont7).strip()
        #     listedCompany_mainBusinessCompositionAnalysis_2[
        #         'listedCompany_mainBusinessCompositionAnalysis_rateOfMargin'] = cont7
        #     listedCompany_mainBusinessCompositionAnalysis_2_time = root11.xpath("./ul/li[2]/a/text()").getall()
        #     listedCompany_mainBusinessCompositionAnalysis_2_time = ''.join(
        #         listedCompany_mainBusinessCompositionAnalysis_2_time).strip()
        #     listedCompany_mainBusinessCompositionAnalysis_2[
        #         'listedCompany_mainBusinessCompositionAnalysis_time'] = listedCompany_mainBusinessCompositionAnalysis_2_time
        #     list2.append(listedCompany_mainBusinessCompositionAnalysis_2)
        # company_ba['listedCompany_mainBusinessCompositionAnalysis_2'] = list2
        #
        # root_trs2 = root.xpath("./div[@class='m_tab_content'][3]/table/tbody/tr")
        # list3 = list()
        # listedCompany_mainBusinessCompositionAnalysis_3 = dict()
        # for root_tr in root_trs2:
        #     cont1 = root_tr.xpath("./td[1]/text()").getall()
        #     cont1 = ''.join(cont1).strip()
        #     listedCompany_mainBusinessCompositionAnalysis_3[
        #         'listedCompany_mainBusinessCompositionAnalysis_bussinessName'] = cont1
        #     cont2 = root_tr.xpath("./td[2]/text()").getall()
        #     cont2 = ''.join(cont2).strip()
        #     listedCompany_mainBusinessCompositionAnalysis_3[
        #         'listedCompany_mainBusinessCompositionAnalysis_operatingIncome'] = cont2
        #     cont3 = root_tr.xpath("./td[3]/text()").getall()
        #     cont3 = ''.join(cont3).strip()
        #     listedCompany_mainBusinessCompositionAnalysis_3[
        #         'listedCompany_mainBusinessCompositionAnalysis_incomeRatio'] = cont3
        #     cont4 = root_tr.xpath("./td[4]/text()").getall()
        #     cont4 = ''.join(cont4).strip()
        #     listedCompany_mainBusinessCompositionAnalysis_3[
        #         'listedCompany_mainBusinessCompositionAnalysis_operatingCosts'] = cont4
        #     cont5 = root_tr.xpath("./td[5]/text()").getall()
        #     cont5 = ''.join(cont5).strip()
        #     listedCompany_mainBusinessCompositionAnalysis_3[
        #         'listedCompany_mainBusinessCompositionAnalysis_costRatio'] = cont5
        #     cont6 = root_tr.xpath("./td[6]/text()").getall()
        #     cont6 = ''.join(cont6).strip()
        #     listedCompany_mainBusinessCompositionAnalysis_3[
        #         'listedCompany_mainBusinessCompositionAnalysis_percentageOfProfit'] = cont6
        #     cont7 = root_tr.xpath("./td[7]/text()").getall()
        #     cont7 = ''.join(cont7).strip()
        #     listedCompany_mainBusinessCompositionAnalysis_3[
        #         'listedCompany_mainBusinessCompositionAnalysis_rateOfMargin'] = cont7
        #     listedCompany_mainBusinessCompositionAnalysis_3_time = root11.xpath("./ul/li[3]/a/text()").getall()
        #     listedCompany_mainBusinessCompositionAnalysis_3_time = ''.join(
        #         listedCompany_mainBusinessCompositionAnalysis_3_time).strip()
        #     listedCompany_mainBusinessCompositionAnalysis_3[
        #         'listedCompany_mainBusinessCompositionAnalysis_time'] = listedCompany_mainBusinessCompositionAnalysis_3_time
        #     list3.append(listedCompany_mainBusinessCompositionAnalysis_3)
        # company_ba['listedCompany_mainBusinessCompositionAnalysis_3'] = list3
        # # dict1['time1']=time1
        # # yewumingcheng = root.xpath("./tbody/tr[1]/td[1]/text()").get()
        # # print(yewumingcheng)
        # # dict1['yewumingcheng'] = yewumingcheng
        # # pingfangmi = root.xpath("./tbody/tr[1]/td[2]/text()").get()
        # # print(pingfangmi)
        # # dict1['pingfangmi'] = pingfangmi
        # # print(dict1)
        # company_ba['listedCompany_mainBusiness'] = listedCompany_mainBusiness
        #
        #
        # root21 = response.xpath("//div[@class='content page_event_content']/div[@class='m_box'][@id='observe']/div[@class='bd']")
        # root211 = root21.xpath("./div[@class='m_tab']/ul/li[1]/a/text()").getall()
        #
        # root212 = root21.xpath("./div[@class='m_tab_content m_tab_content2']/p[2]/text()").getall()
        # root212 = ''.join(root212).strip()
        # print(root212)
        # print(" ".join(str(i) for i in root212))


        # 去除列表中多余字符串以及换行
        # lists = [x.strip() for x in root212]
        # set = list(set(lists))
        # set.sort(key=lists.index)
        # set.remove('')
        # print(set)

        # 董事会经营评述 模块 2020-08-10 -- myh
        listedCompany_businessAnalysis_reviewOfBoardOperation = list()
        # 日期列表
        date_list = response.xpath("//div[@id='observe']//div[@class='m_tab']//a/text()").getall()
        # 日期对应的p列表
        p_list = response.xpath("//p[@class='f14 none clearfix pr']")
        print(len(p_list))
        for i in range(0, len(date_list)):
            content_dict = dict()
            # 1.日期
            content_dict['listedCompany_businessAnalysis_reviewOfBoardOperation_date'] = date_list[i]
            # 2.内容
            content_list = p_list[i].xpath("./text()").getall()
            # 去空格等空字符
            content_list = [content.replace("\r", "").strip() for content in content_list]
            content_dict['listedCompany_businessAnalysis_reviewOfBoardOperation_content'] = "".join(content_list)
            # 将字典对象加入该模块的列表
            listedCompany_businessAnalysis_reviewOfBoardOperation.append(content_dict)
        # 将该列表 传递到item对应的字段
        company_ba['listedCompany_businessAnalysis_reviewOfBoardOperation'] = listedCompany_businessAnalysis_reviewOfBoardOperation

        yield company_ba

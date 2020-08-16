# -*- coding: utf-8 -*-
import random
import time
from jqka_new.items import JqkaNewItem
from scrapy.http import Request
# 获取excel需要引入的包
from openpyxl import load_workbook
from selenium import webdriver
import scrapy


class JqkaNewSpderSpider(scrapy.Spider):
    name = 'jqka_new_spder'
    allowed_domains = ['http://basic.10jqka.com.cn/000056/news.html']
    start_urls = ['http://http://basic.10jqka.com.cn/000056/news.html/']
    def start_requests(self):
        yield Request("http://basic.10jqka.com.cn/603221/news.html",
                      headers={
                          'User-Agent': "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.87 Safari/537.36"})


    def parse(self, response):
        book = load_workbook(filename=r"C:\python\lnuSpider\data\exel\com_list.xlsx")
        sheet = book.active
        data = []
        data1 = []
        data2 = []
        data3 = []
        data4 = []
        row_num = 1
        while row_num <= 1:
            # 将表中第一列的1-100行数据写入data数组中
            data.append(sheet.cell(row=row_num, column=3).value)
            data1.append(sheet.cell(row=row_num, column=1).value)
            data3.append(sheet.cell(row=row_num, column=2).value)
            data4.append(sheet.cell(row=row_num, column=4).value)
            data2.append(row_num)
            row_num = row_num + 1
        for i in data2:
            # url = 'http://basic.10jqka.com.cn/'+data[i]+'/company.html'
            # print(data[i-1])
            a = str(data[i - 1])
            print(type(a))
            listedCompany_url = 'http://basic.10jqka.com.cn/' + a + '/news.html'
            company_news = JqkaNewItem()
            company_news['listedCompany_url'] = listedCompany_url
            listedCompany_id = data1[i - 1]
            company_news['listedCompany_id'] = listedCompany_id
            listedCompany_name = data3[i - 1]
            company_news['listedCompany_name'] = listedCompany_name
            listedCompany_fullName = data4[i - 1]
            company_news['listedCompany_fullName'] = listedCompany_fullName
            # print(listedCompany_id)
            yield scrapy.Request(company_news['listedCompany_url'],
                                 meta={'company_news': company_news}, callback=self.detail_ni, dont_filter=True)
        return

    def detail_ni(self, response):
        company_news = response.meta['company_news']
        # root = response.xpath("//div[@class='content page_event_content']"
        #                       "/div[@class='m_box subnews'][@id='pub'][@stat='pub_pub']"
        #                       "/div[@class='bd']/div[@class='m_tab_content']/*").getall()
        # print(root)
        # root1 = response.xpath("//div[@class='m_dlbox'][@id='pull_all']/dl")
        # print(type(root1))
        # print(root1)

        options = webdriver.FirefoxOptions()
        options.add_argument("--headless")  # 设置火狐为headless无界面模式
        options.add_argument("--disable-gpu")

        driver_path = r"D:\project\geckodriver.exe"
        driver = webdriver.Firefox(executable_path=driver_path, firefox_options=options)
        driver.get(company_news['listedCompany_url'])

        while True:
            # c = driver.find_element_by_xpath("//div[@class='m_dlbox'][@id='pull_all']")
            # print(c.get_attribute('innerHTML'))
            root1 = driver.find_elements_by_xpath("//div[@class='m_dlbox'][@id='pull_all']//dl")

            # root1.get_attribute('innerHTML')
            # print(root1)
            for root2 in root1:
                news_date = root2.find_element_by_xpath("./dt/span[@class='date']")
                # print(news_date.text)
                news_date = news_date.text
                news_tag = root2.find_element_by_xpath("./dt/span[@class='title']/a/strong")
                # print(news_tag.text)
                news_tag = news_tag.text
                news_url = root2.find_element_by_xpath("./dt/span[@class='title']/a")
                # print(news_url.get_attribute("href"))
                # print(news_tag)
                # print(news_url)
                # company_news['news_url'] = news_url.get_attribute("href")
                yield scrapy.Request(news_url.get_attribute("href"),
                                     meta={"news_url": news_url.get_attribute("href"), "url": company_news['listedCompany_url']
                                         , "news_date": news_date, "news_tag": news_tag, 'company_news': company_news},
                                     callback=self.detail_ni1,
                                     dont_filter=True)

            ul = driver.find_element_by_css_selector("[class='splpager clearfix light-theme simple-pagination']")
            li_label_s = ul.find_elements_by_xpath('./ul/li')
            # print("测试 'innerHTML ："+li_label_s[-1].get_attribute('innerHTML'))
            xia_yi_ye = li_label_s[-1].get_attribute('innerHTML')
            if "下一页</a>" in xia_yi_ye:
                print("是a!")
                element = driver.find_element_by_css_selector("[class='page-link next']")
                element.click()
                time.sleep(3)
            elif "下一页</span>" in xia_yi_ye:
                print("是span!")
                break

    def detail_ni1(self, response):
        print("======进入内页======")
        company_news = JqkaNewItem()
        company_news = response.meta['company_news']
        # company_news['listedCompany_news_announceList_url0'] = response.meta['news_url']
        # print(company_news['news_url0'])
        # company_news['listedCompany_news_announceList_url'] = response.meta['url']
        company_news['listedCompany_news_announceList_tag'] = response.meta['news_tag']
        company_news['listedCompany_news_announceList_date'] = response.meta['news_date']
        time.sleep(random.randint(1, 2))
        yield company_news

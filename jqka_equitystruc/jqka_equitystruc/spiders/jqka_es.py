# -*- coding: utf-8 -*-
import random

import requests
import scrapy
from jqka_equitystruc.items import JqkaEquitystrucItem
from lxml import etree
from scrapy import item
from scrapy.http import Request
# 获取excel需要引入的包
from openpyxl import load_workbook
from selenium import webdriver
import time
import scrapy
import pandas as pd


class JqkaEsSpider(scrapy.Spider):
    name = 'jqka_es'
    allowed_domains = ['http://basic.10jqka.com.cn/603221/holder.html']
    start_urls = ['http://http://basic.10jqka.com.cn/603221/holder.html/']

    # start_urls = ['http://basic.10jqka.com.cn/603221/operate.html/']

    def start_requests(self):
        yield Request("http://basic.10jqka.com.cn/603221/holder.html", headers={
            'User-Agent': "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.87 Safari/537.36"})

    def parse(self, response):
        book = load_workbook(filename=r"C:\python\lnuSpider\data\exel\com_list.xlsx")
        sheet = book.active
        data = []
        data1 = []
        data2 = []
        data3 = []
        row_num = 1
        while row_num <= 1:
            # 将表中第一列的1-100行数据写入data数组中
            data.append(sheet.cell(row=row_num, column=3).value)
            data1.append(sheet.cell(row=row_num, column=1).value)
            data3.append(sheet.cell(row=row_num, column=2).value)
            data2.append(row_num)
            row_num = row_num + 1
        for i in data2:
            # url = 'http://basic.10jqka.com.cn/'+data[i]+'/company.html'
            # print(data[i-1])
            listedCompany_url = 'http://basic.10jqka.com.cn/' + data[i - 1] + '/holder.html'
            company_es = JqkaEquitystrucItem()
            company_es['listedCompany_url'] = listedCompany_url
            listedCompany_id = data1[i - 1]
            company_es['listedCompany_id'] = listedCompany_id
            listedCompany_name = data3[i - 1]
            company_es['listedCompany_name'] = listedCompany_name
            # print(listedCompany_id)
            # pandas读取表格
            # res_elements = etree.HTML(response.text)
            # table = res_elements.xpath("//table[@class='tbody']")
            # table = etree.tostring(table[0], encoding='utf-8').decode()
            # df = pd.read_html(table, encoding='utf-8', header=0)
            # results = list(df.T.to_dict().values())  # 转换成列表嵌套字典的格式
            # print(results)
            yield scrapy.Request(company_es['listedCompany_url'],
                                 meta={'company_es': company_es}, callback=self.detail_ni, dont_filter=True)
        return

    def detail_ni(self, response):
        company_es = response.meta['company_es']
        # root = response.xpath(".//div[@id='gdrsTable'][@class='holernumcomp gdrs_table']//div[@class='table_data']")
        item = response.xpath("//div[@id='bd_0']/div[1]/ul/li")
        # print(item)
        listedCompany_shareholderResearch_topTenShareholders = list()
        list1 = list()
        # print(len(itm))
        for x in range(len(item)):
            list0 = list()
            x = x + 1
            print(x)
            root2 = response.xpath("//div[@id='bd_0']/div[{}]/table".format(x+1))
            root3 = root2.xpath("./tbody/tr")
            for rot3 in root3:
                single=dict()

                listedCompany_shareholderResearch_topTenShareholders_shareholderName = rot3.xpath("./th/a/text()").getall()
                listedCompany_shareholderResearch_topTenShareholders_shareholderName = ''.join(listedCompany_shareholderResearch_topTenShareholders_shareholderName).strip()
                single['listedCompany_shareholderResearch_topTenShareholders_shareholderName'] = listedCompany_shareholderResearch_topTenShareholders_shareholderName

                listedCompany_shareholderResearch_topTenShareholders_shareType = rot3.xpath("./td[5]/text()").getall()
                listedCompany_shareholderResearch_topTenShareholders_shareType = ''.join(
                    listedCompany_shareholderResearch_topTenShareholders_shareType).strip()
                single['listedCompany_shareholderResearch_topTenShareholders_shareType'] = listedCompany_shareholderResearch_topTenShareholders_shareType

                listedCompany_shareholderResearch_topTenShareholders_sharesHoldNumber = rot3.xpath("./td[1]/text()").getall()
                listedCompany_shareholderResearch_topTenShareholders_sharesHoldNumber = ''.join(
                    listedCompany_shareholderResearch_topTenShareholders_sharesHoldNumber).strip()
                single['listedCompany_shareholderResearch_topTenShareholders_sharesHoldNumber'] = listedCompany_shareholderResearch_topTenShareholders_sharesHoldNumber

                listedCompany_shareholderResearch_topTenShareholders_shareHoldingChange = rot3.xpath("./td[2]/text()").getall()
                listedCompany_shareholderResearch_topTenShareholders_shareHoldingChange = ''.join(
                    listedCompany_shareholderResearch_topTenShareholders_shareHoldingChange).strip()
                single['listedCompany_shareholderResearch_topTenShareholders_shareHoldingChange'] = listedCompany_shareholderResearch_topTenShareholders_shareHoldingChange

                listedCompany_shareholderResearch_topTenShareholders_totalEquityProportion = rot3.xpath("./td[3]/text()").getall()
                listedCompany_shareholderResearch_topTenShareholders_totalEquityProportion = ''.join(
                    listedCompany_shareholderResearch_topTenShareholders_totalEquityProportion).strip()
                single['listedCompany_shareholderResearch_topTenShareholders_totalEquityProportion'] = listedCompany_shareholderResearch_topTenShareholders_totalEquityProportion

                listedCompany_shareholderResearch_topTenShareholders_increaseOrDecrease = rot3.xpath("./td[4]/text()").getall()
                listedCompany_shareholderResearch_topTenShareholders_increaseOrDecrease = ''.join(
                    listedCompany_shareholderResearch_topTenShareholders_increaseOrDecrease).strip()
                single['listedCompany_shareholderResearch_topTenShareholders_increaseOrDecrease'] = listedCompany_shareholderResearch_topTenShareholders_increaseOrDecrease
                list0.append(single)
            list1.append(list0)
        listedCompany_shareholderResearch_topTenShareholders.append(list1)
        company_es[
            'listedCompany_shareholderResearch_topTenShareholders'] = listedCompany_shareholderResearch_topTenShareholders
        # print(list2)

        root = response.xpath("//div[@class='content page_event_content']//div[@id='gdrsTable'][@class='holernumcomp gdrs_table']")
        # print(root)
        root1 = root.xpath("./div[@class='scroll_container']/div[@class='table_data']/div[@class='data_tbody']")
        itm = root1.xpath("./table[@class='top_thead']/tr/th")
        listedCompany_shareholderResearch_shareholderNum_time = list()
        for itm1 in itm:
            listedCompany_shareholderResearch_shareholderNum_tim = itm1.xpath("./div/text()").getall()
            listedCompany_shareholderResearch_shareholderNum_tim = ''.join(listedCompany_shareholderResearch_shareholderNum_tim).strip()
            listedCompany_shareholderResearch_shareholderNum_time.append(listedCompany_shareholderResearch_shareholderNum_tim)
        company_es['listedCompany_shareholderResearch_shareholderNum_time'] = listedCompany_shareholderResearch_shareholderNum_time

        listedCompany_shareholderResearch_shareholderNum_totalShareholdersNumber = list()
        root2 = root1.xpath("./table[@class='tbody']/tr[1]/td")
        for root3 in root2:
            listedCompany_shareholderResearch_shareholderNum_totalShareholdersNumbe = root3.xpath("./div/text()").getall()
            listedCompany_shareholderResearch_shareholderNum_totalShareholdersNumbe = ''.join(
                listedCompany_shareholderResearch_shareholderNum_totalShareholdersNumbe).strip()
            listedCompany_shareholderResearch_shareholderNum_totalShareholdersNumber.append(
                listedCompany_shareholderResearch_shareholderNum_totalShareholdersNumbe)
        company_es[
            'listedCompany_shareholderResearch_shareholderNum_totalShareholdersNumber'] = listedCompany_shareholderResearch_shareholderNum_totalShareholdersNumber

        listedCompany_shareholderResearch_shareholderNum_comparedPreviousPeriodChange = list()
        root2 = root1.xpath("./table[@class='tbody']/tr[2]/td")
        for root3 in root2:
            listedCompany_shareholderResearch_shareholderNum_comparedPreviousPeriodChang = root3.xpath("./span/text()").getall()
            listedCompany_shareholderResearch_shareholderNum_comparedPreviousPeriodChang = ''.join(
                listedCompany_shareholderResearch_shareholderNum_comparedPreviousPeriodChang).strip()
            listedCompany_shareholderResearch_shareholderNum_comparedPreviousPeriodChange.append(
                listedCompany_shareholderResearch_shareholderNum_comparedPreviousPeriodChang)
        company_es[
            'listedCompany_shareholderResearch_shareholderNum_comparedPreviousPeriodChange'] = listedCompany_shareholderResearch_shareholderNum_comparedPreviousPeriodChange

        listedCompany_shareholderResearch_shareholderNum_perCapitaCirculatingShares = list()
        root2 = root1.xpath("./table[@class='tbody']/tr[3]/td")
        for root3 in root2:
            listedCompany_shareholderResearch_shareholderNum_perCapitaCirculatingShare = root3.xpath("./text()").getall()
            listedCompany_shareholderResearch_shareholderNum_perCapitaCirculatingShare = ''.join(
                listedCompany_shareholderResearch_shareholderNum_perCapitaCirculatingShare).strip()
            listedCompany_shareholderResearch_shareholderNum_perCapitaCirculatingShares.append(
                listedCompany_shareholderResearch_shareholderNum_perCapitaCirculatingShare)
        company_es[
            'listedCompany_shareholderResearch_shareholderNum_perCapitaCirculatingShares'] = listedCompany_shareholderResearch_shareholderNum_perCapitaCirculatingShares

        listedCompany_shareholderResearch_shareholderNum_perCapitaCirculationChanges = list()
        root2 = root1.xpath("./table[@class='tbody']/tr[4]/td")
        for root3 in root2:
            listedCompany_shareholderResearch_shareholderNum_perCapitaCirculationChange = root3.xpath("./span/text()").getall()
            listedCompany_shareholderResearch_shareholderNum_perCapitaCirculationChange = ''.join(
                listedCompany_shareholderResearch_shareholderNum_perCapitaCirculationChange).strip()
            listedCompany_shareholderResearch_shareholderNum_perCapitaCirculationChanges.append(
                listedCompany_shareholderResearch_shareholderNum_perCapitaCirculationChange)
        company_es[
            'listedCompany_shareholderResearch_shareholderNum_perCapitaCirculationChanges'] = listedCompany_shareholderResearch_shareholderNum_perCapitaCirculationChanges

        listedCompany_shareholderResearch_shareholderNum_industryAverage = list()
        root2 = root1.xpath("./table[@class='tbody']/tr[5]/td")
        for root3 in root2:
            listedCompany_shareholderResearch_shareholderNum_industryAverag = root3.xpath("./span/text()").getall()
            listedCompany_shareholderResearch_shareholderNum_industryAverag = ''.join(listedCompany_shareholderResearch_shareholderNum_industryAverag).strip()
            listedCompany_shareholderResearch_shareholderNum_industryAverage.append(listedCompany_shareholderResearch_shareholderNum_industryAverag)
        company_es['listedCompany_shareholderResearch_shareholderNum_industryAverage'] = listedCompany_shareholderResearch_shareholderNum_industryAverage



        root_1 = response.xpath(
            "//div[@class='content page_event_content']//div[@id='flowholder'][@class='m_box hold_detail z100 gssj_scroll gssj_scroll1']")
        root_11 = root_1.xpath("//div[@id='bd_1']/div[@class='m_tab mt15']/ul/li")
        listedCompany_shareholderResearch_topTenCurrentShareholders_time = list()
        for root_111 in root_11:
            listedCompany_shareholderResearch_topTenCurrentShareholders_tim = root_111.xpath("./a/text()").getall()
            listedCompany_shareholderResearch_topTenCurrentShareholders_tim = ''.join(listedCompany_shareholderResearch_topTenCurrentShareholders_tim).strip()
            listedCompany_shareholderResearch_topTenCurrentShareholders_time.append(listedCompany_shareholderResearch_topTenCurrentShareholders_tim)
        company_es['listedCompany_shareholderResearch_topTenCurrentShareholders_time'] = listedCompany_shareholderResearch_topTenCurrentShareholders_time

        listedCompany_shareholderResearch_topTenCurrentShareholders = list()
        root_12 = root_1.xpath("//div[@id='bd_1']/div[@class='m_tab_content2 clearfix']/table/tbody[1]/tr")
        for root_121 in root_12:
            single5 = dict()
            content_5 = root_121.xpath("./th/a/text()").getall()
            single5['listedCompany_shareholderResearch_topTenCurrentShareholders_shareholderName'] = "".join(content_5).strip()
            content_6 = root_121.xpath("./td[1]/text()").getall()
            single5['listedCompany_shareholderResearch_topTenCurrentShareholders_sharesHoldNumber'] = "".join(content_6).strip()
            content_7 = root_121.xpath("./td[2]/span/text()").getall()
            single5['listedCompany_shareholderResearch_topTenCurrentShareholders_shareHoldChange'] = "".join(content_7).strip()
            content_8 = root_121.xpath("./td[3]/text()").getall()
            single5['listedCompany_shareholderResearch_topTenCurrentShareholders_totalEquityProportion'] = "".join(content_8).strip()
            content_9 = root_121.xpath("./td[5]/text()").getall()
            single5['listedCompany_shareholderResearch_topTenCurrentShareholders_shareType'] = "".join(content_9).strip()
            listedCompany_shareholderResearch_topTenCurrentShareholders.append(single5)
        company_es['listedCompany_shareholderResearch_topTenCurrentShareholders'] = listedCompany_shareholderResearch_topTenCurrentShareholders

        listedCompany_shareholderResearch_topTenCurrentShareholders_increaseOrDecrease = root_1.xpath(
            "//div[@id='bd_1']/div[@class='m_tab_content2 clearfix']/table/caption//text()").getall()
        listedCompany_shareholderResearch_topTenCurrentShareholders_increaseOrDecrease = "".join(
            listedCompany_shareholderResearch_topTenCurrentShareholders_increaseOrDecrease).strip().replace('\t', '')
        company_es[
            'listedCompany_shareholderResearch_topTenCurrentShareholders_increaseOrDecrease'] = listedCompany_shareholderResearch_topTenCurrentShareholders_increaseOrDecrease

        # # pandas读取表格
        # res_elements = etree.HTML(response.text)
        # table = res_elements.xpath("//table[@class='m_table m_hl ggintro']")
        # table = etree.tostring(table[2], encoding='utf-8').decode()
        # df = pd.read_html(table, encoding='utf-8', header=0)[0]
        # results = list(df.T.to_dict().values())  # 转换成列表嵌套字典的格式
        # print(results)
        time.sleep(random.randint(3, 6))
        yield company_es

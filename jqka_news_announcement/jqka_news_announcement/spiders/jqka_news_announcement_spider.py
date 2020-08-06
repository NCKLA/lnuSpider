# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import time
import scrapy
from jqka_news_announcement.items import JqkaNewsAnnouncementItem
from selenium import webdriver
from selenium.webdriver.firefox.options import Options
# from selenium.webdriver.chrome.options import Options


class JqkaNewsAnnouncementSpiderSpider(scrapy.Spider):
    name = 'jqka_news_announcement_spider'
    #allowed_domains = ['basic.10jqka.com']
    start_urls = ['http://basic.10jqka.com.cn/603221/news.html']

    def __init__(self):
        options = Options()
        options.headless = True
        driver_path = r"D:\project\geckodriver.exe"
        self.chrome = webdriver.Firefox(executable_path=driver_path, firefox_options=options)
        self.chrome_2 = webdriver.Firefox(executable_path=driver_path, firefox_options=options)
        self.chrome.set_page_load_timeout(40)
        self.chrome_2.set_page_load_timeout(40)

    # 读取公司对应url 的函数
    def parse(self, response):
        # 读取excel表格
        book = load_workbook(filename=r"C:\python\lnuSpider\data\exel\com_list.xlsx")
        sheet = book.active
        data = []
        data1 = []
        data2 = []
        data3 = []
        row_num = 1
        while row_num <= 3815:
            # 将表中第一列的1-100行数据写入data数组中
            data.append(sheet.cell(row=row_num, column=3).value)
            data1.append(sheet.cell(row=row_num, column=1).value)
            data3.append(sheet.cell(row=row_num, column=2).value)
            data2.append(row_num)
            row_num = row_num + 1
        for i in data2:
            a = str(data[i - 1])
            listedCompany_url = 'http://basic.10jqka.com.cn/' + a + '/news.html'
            company_na = JqkaNewsAnnouncementItem()
            company_na['listedCompany_url'] = listedCompany_url
            listedCompany_id = data1[i - 1]
            company_na['listedCompany_id'] = listedCompany_id
            listedCompany_name = data3[i - 1]
            company_na['listedCompany_name'] = listedCompany_name
            yield scrapy.Request(company_na['listedCompany_url'], meta={'company_na': company_na}, callback=self.parse_outpage, dont_filter=True)


    # 1.外页解析函数
    def parse_outpage(self, response):
        # 获得相应公司对应的公司信息
        company_na = response.meta['company_na']
        item = JqkaNewsAnnouncementItem()
        # 公司信息
        item['listedCompany_id'] = company_na['listedCompany_id']
        item['listedCompany_name'] = company_na['listedCompany_name']
        item['listedCompany_url'] = company_na['listedCompany_url']
        # 动态新闻 模块
        item['listedCompany_news_hotnews'] = list()
        # 由于该模块的信息是动态加载的，所以只有通过selenium才能获取到动态加载后的数据
        self.chrome.get(item['listedCompany_url'])
        time.sleep(1)

        while True:
            # 这个必须写到while里面，因为若不写到里面翻页之后便找不到元素，实现一种刷新的效果
            news_dts = self.chrome.find_elements_by_xpath("//div[@id='mine']/div[@class='bd']//dt")
            for news_dt in news_dts:
                news_dict = dict()
                news_dict['listedCompany_news_hotnews_tag'] = news_dt.find_element_by_xpath(".//strong").text.strip()
                news_dict['listedCompany_news_hotnews_time'] = news_dt.find_element_by_xpath("./span[last()]").text.strip()

                # 通过  parse_innerpage方法 传递 news_dict对象 从而进行完善字典对象
                inner_url = news_dt.find_element_by_xpath("./span[1]/a").get_attribute("href").strip()
                news_dict = self.parse_innerpage(inner_url, news_dict)
                # 将字典对象加入列表
                item['listedCompany_news_hotnews'].append(news_dict)

            # 下一页按钮 对象
            next_ele = self.chrome.find_element_by_xpath("//div[@id='mine']//div[@class='m_page main_page']/a[last()]")
            # 是最后一页 则跳出循环
            if next_ele.get_attribute("class") == "disable":
                break
            # 不是最后一页 则翻页
            else:
                next_ele.click()
                #time.sleep(1)

        # 传递到管道进行存储
        yield item



    # 2.内页解析函数
    def parse_innerpage(self, inner_url, news_dict):
        # 因为实际的内页url和源码内页url不同，所以要借助selenium模拟跳转内页后再获取current_url
        # 并且内页数据为动态加载的数据，也需要借助selenium来获取
        self.chrome_2.get(inner_url)
        time.sleep(1)
        # 因为内页的源码不同，所以根据内页的url是否包含某些字符串来判断源码并且根据源码写xpath
        # 因为内页的源码不同，所以根据内页的url是否包含某些字符串来判断源码并且根据源码写xpath
        # 因为内页的源码不同，所以根据内页的url是否包含某些字符串来判断源码并且根据源码写xpath
        # 第一种：^stock*  | ^goodsfu*  |  ^news*
        if ("stock" in self.chrome_2.current_url) or ("goodsfu" in self.chrome_2.current_url):
            # 标题
            if self.chrome_2.find_elements_by_xpath("//div[@class='main-fl fl']/h2"):
                news_dict['listedCompany_news_hotnews_title'] = self.chrome_2.find_elements_by_xpath("//div[@class='main-fl fl']/h2")[0].text.strip()
            else:
                news_dict['listedCompany_news_hotnews_title'] = '暂无内容'
            # 来源
            if self.chrome_2.find_elements_by_xpath("//div[@class='main-fl fl']//span[@id='source_baidu']"):
                news_dict['listedCompany_news_hotnews_refer'] = self.chrome_2.find_elements_by_xpath("//div[@class='main-fl fl']//span[@id='source_baidu']")[0].text.strip()
            else:
                news_dict['listedCompany_news_hotnews_refer'] = '暂无内容'
            # 内容
            p_list = self.chrome_2.find_elements_by_xpath("//div[@class='main-text atc-content']//p")
            content_list = [i.text.strip() for i in p_list]
            news_dict['listedCompany_news_hotnews_content'] = "".join(content_list).strip()

        # 第二种：^research*
        elif "research" in self.chrome_2.current_url:
            if self.chrome_2.find_elements_by_xpath("//div[@class='main-fl fl']/h2"):
                news_dict['listedCompany_news_hotnews_title'] = self.chrome_2.find_elements_by_xpath("//div[@class='main-fl fl']/h2")[0].text.strip()
            else:
                news_dict['listedCompany_news_hotnews_title'] = "暂无内容"
            if self.chrome_2.find_elements_by_xpath("//div[@class='main-fl fl']//span[@id='source_baidu']"):
                news_dict['listedCompany_news_hotnews_refer'] = self.chrome_2.find_elements_by_xpath("//div[@class='main-fl fl']//span[@id='source_baidu']")[0].text.strip()
            else:
                news_dict['listedCompany_news_hotnews_refer'] = "暂无内容"
            p_list = self.chrome_2.find_elements_by_xpath("//div[@class='main-text atc-content']//p")
            content_list = [i.text.strip() for i in p_list]
            news_dict['listedCompany_news_hotnews_content'] = "".join(content_list).strip()

        # 第三种： ^jiaju*
        elif "jiaju" in self.chrome_2.current_url:
            if self.chrome_2.find_elements_by_xpath("//div[@class='main-left fl']/h1"):
                news_dict['listedCompany_news_hotnews_title'] = self.chrome_2.find_elements_by_xpath("//div[@class='main-left fl']/h1")[0].text.strip()
            else:
                news_dict['listedCompany_news_hotnews_title'] = "暂无内容"
            news_dict['listedCompany_news_hotnews_refer'] = "暂无内容"
            p_list = self.chrome_2.find_elements_by_xpath("//div[@id='articleText']//p")
            content_list = [i.text.strip() for i in p_list]
            news_dict['listedCompany_news_hotnews_content'] = "".join(content_list).strip()

        # 第四种： ^bjnews*
        elif "bjnews" in self.chrome_2.current_url:
            if self.chrome_2.find_elements_by_xpath("//div[@class='title']/h1"):
                news_dict['listedCompany_news_hotnews_title'] = self.chrome_2.find_elements_by_xpath("//div[@class='title']/h1")[0].text.strip()
            else:
                news_dict['listedCompany_news_hotnews_title'] = "暂无内容"
            news_dict['listedCompany_news_hotnews_refer'] = "暂无内容"
            p_list = self.chrome_2.find_elements_by_xpath("//div[@class='content']/p")
            content_list = [i.text.strip() for i in p_list]
            news_dict['listedCompany_news_hotnews_content'] = "".join(content_list).strip()

        # 第五种： ^egsea*
        elif "egsea" in self.chrome_2.current_url:
            if self.chrome_2.find_elements_by_xpath("//div[@class='article-title']/h1"):
                news_dict['listedCompany_news_hotnews_title'] = self.chrome_2.find_elements_by_xpath("//div[@class='article-title']/h1")[0].text
            else:
                news_dict['listedCompany_news_hotnews_title'] = "暂无内容"
            if self.chrome_2.find_elements_by_xpath("//span[@class='source']"):
                news_dict['listedCompany_news_hotnews_refer'] = self.chrome_2.find_elements_by_xpath("//span[@class='source']")[0].text.strip()
            else:
                news_dict['listedCompany_news_hotnews_refer'] = "暂无内容"
            p_list = self.chrome_2.find_elements_by_xpath("//div[@class='article-content-detail']//p")
            content_list = [i.text.strip() for i in p_list]
            news_dict['listedCompany_news_hotnews_content'] = "".join(content_list).strip()

        # 第六种： ^weixin*
        elif "weixin" in self.chrome_2.current_url:
            if self.chrome_2.find_elements_by_xpath("//h2[@class='rich_media_title']"):
                news_dict['listedCompany_news_hotnews_title'] = self.chrome_2.find_elements_by_xpath("//h2[@class='rich_media_title']")[0].text
            else:
                news_dict['listedCompany_news_hotnews_title'] = "暂无内容"
            if self.chrome_2.find_elements_by_xpath("//a[@id='js_name']"):
                news_dict['listedCompany_news_hotnews_refer'] = self.chrome_2.find_elements_by_xpath("//a[@id='js_name']")[0].text
            else:
                news_dict['listedCompany_news_hotnews_refer'] = "暂无内容"
            p_list = self.chrome_2.find_elements_by_xpath("//div[@id='js_content']//p")
            content_list = [i.text.strip() for i in p_list]
            news_dict['listedCompany_news_hotnews_content'] = "".join(content_list).strip()

        # 第七种类： ^jiemian*
        elif "jiemian" in self.chrome_2.current_url:
            if self.chrome_2.find_elements_by_xpath("//div[@class='article-header']/h1"):
                news_dict['listedCompany_news_hotnews_title'] = self.chrome_2.find_elements_by_xpath("//div[@class='article-header']/h1")[0].text
            else:
                news_dict['listedCompany_news_hotnews_title'] = "暂无内容"
            if self.chrome_2.find_elements_by_xpath("//div[@class='article-info']//span[4]"):
                news_dict['listedCompany_news_hotnews_refer'] = self.chrome_2.find_elements_by_xpath("//div[@class='article-info']//span[4]")[0].text
            else:
                news_dict['listedCompany_news_hotnews_refer'] = "暂无内容"
            p_list = self.chrome_2.find_elements_by_xpath("//div[@class='article-content']//p")
            content_list = [i.text.strip() for i in p_list]
            news_dict['listedCompany_news_hotnews_content'] = "".join(content_list).strip()

        # 第八种： ^nbd*
        elif "nbd" in self.chrome_2.current_url:
            if self.chrome_2.find_elements_by_xpath("//div[@class='g-article-top']/h1"):
                news_dict['listedCompany_news_hotnews_title'] = self.chrome_2.find_elements_by_xpath("//div[@class='g-article-top']/h1")[0].text
            else:
                news_dict['listedCompany_news_hotnews_title'] = "暂无内容"
            if self.chrome_2.find_elements_by_xpath("//span[@class='source']"):
                news_dict['listedCompany_news_hotnews_refer'] = self.chrome_2.find_elements_by_xpath("//span[@class='source']")[0].text
            else:
                news_dict['listedCompany_news_hotnews_refer'] = "暂无内容"
            p_list = self.chrome_2.find_elements_by_xpath("//div[@class='g-articl-text']/p")
            content_list = [i.text.strip() for i in p_list]
            news_dict['listedCompany_news_hotnews_content'] = "".join(content_list).strip()

        else:
            news_dict['listedCompany_news_hotnews_title'] = "暂无内容"
            news_dict['listedCompany_news_hotnews_refer'] = "暂无内容"
            news_dict['listedCompany_news_hotnews_content'] = "暂无内容"

        return news_dict
# -*- coding: utf-8 -*-
import random
import re
from scrapy import item
from scrapy.http import Request
from openpyxl import load_workbook
from selenium import webdriver
import time
import scrapy
from jqka_company_event.items import JqkaCompanyEventItem


class JqkaCompanyEventSpiderSpider(scrapy.Spider):
    name = 'jqka_company_event_spider'
    allowed_domains = ['basic.10.jqka.com']
    start_urls = ['http://basic.10jqka.com.cn/603221/event.html']

    def parse(self, response):
        # 读取excel表格
        book = load_workbook(filename=r"C:\python\lnuSpider\data\exel\com_list.xlsx")
        sheet = book.active
        data = []
        data1 = []
        data2 = []
        data3 = []
        data4 = []
        row_num = 1
        while row_num <= 3815:
            # 将表中第一列的1-100行数据写入data数组中
            data.append(sheet.cell(row=row_num, column=3).value)
            data1.append(sheet.cell(row=row_num, column=1).value)
            data3.append(sheet.cell(row=row_num, column=2).value)
            data4.append(sheet.cell(row=row_num, column=4).value)
            data2.append(row_num)
            row_num = row_num + 1
        for i in data2:
            a = str(data[i - 1])
            listedCompany_url = 'http://basic.10jqka.com.cn/' + a + '/event.html'
            company_ce = JqkaCompanyEventItem()
            company_ce['listedCompany_url'] = listedCompany_url
            listedCompany_id = data1[i - 1]
            company_ce['listedCompany_id'] = listedCompany_id
            listedCompany_name = data3[i - 1]
            company_ce['listedCompany_name'] = listedCompany_name
            listedCompany_fullName = data4[i - 1]
            company_ce['listedCompany_fullName'] = listedCompany_fullName
            yield scrapy.Request(company_ce['listedCompany_url'], meta={'company_ce': company_ce},callback=self.detail_ni, dont_filter=True)


    def detail_ni(self, response):
        # 获得相应公司对应的公司信息
        company_ce = response.meta['company_ce']
        item = JqkaCompanyEventItem()

        # 公司信息
        item['listedCompany_id'] = company_ce['listedCompany_id']
        item['listedCompany_name'] = company_ce['listedCompany_name']
        item['listedCompany_url'] = company_ce['listedCompany_url']
        item['listedCompany_fullName'] = company_ce['listedCompany_fullName']

        # 1.违规处理 模块
        # listedCompany_companyEvent_handofViolation
        # 首先判断该公司有无该模块
        if response.xpath("//div[@id='violate']"):
            item['listedCompany_companyEvent_handofViolation'] = list()
            violate_list = response.xpath("//table[@class='mt15 m_table']")
            for violate in violate_list:
                violate_dict = dict()
                violate_dict['listedCompany_companyEvent_handofViolation_announcementDate'] = violate.xpath(".//tr[1]/td[1]/text()").get()
                violate_dict['listedCompany_companyEvent_handofViolation_punishmentObject '] = violate.xpath(".//tr[3]/td[1]/text()[2]").get().strip()
                violate_dict['listedCompany_companyEvent_handofViolation_punishmentTypes '] = violate.xpath(".//tr[1]/td[3]/text()").get()
                violate_dict['listedCompany_companyEvent_handofViolation_penaltyAmount'] = violate.xpath(".//tr[1]/td[2]/text()").get()
                violate_dict['listedCompany_companyEvent_handofViolation_dealWithPeople'] = violate.xpath(".//tr[2]/td/text()[2]").get().strip()
                violate_dict['listedCompany_companyEvent_handofViolation_irregularities'] = violate.xpath(".//tr[3]/td[2]/a/@content").get().strip()
                violate_dict['listedCompany_companyEvent_handofViolation_disciplinaryMeasures'] = violate.xpath(".//tr[4]/td/p/text()").get().strip()
                # 加入列表
                item['listedCompany_companyEvent_handofViolation'].append(violate_dict)
        else:
            item['listedCompany_companyEvent_handofViolation'] = []


        # 2.高管持股变动 模块
        # listedCompany_companyEvent_executiveShareholdingchanges
        # 首先判断该公司有无该模块
        if response.xpath("//div[@id='manager']"):
            item['listedCompany_companyEvent_executiveShareholdingchanges'] = list()
            tr_list= response.xpath("//div[@id='manager']//tbody/tr")
            for tr in tr_list:
                manager_dict = dict()
                manager_dict['listedCompany_companyEvent_executiveShareholdingchanges_changeDate'] = tr.xpath('./td[1]/text()').get()
                manager_dict['listedCompany_companyEvent_executiveShareholdingchanges_variablePerson'] = tr.xpath('./td[2]/text()').get()
                manager_dict['listedCompany_companyEvent_executiveShareholdingchanges_relationshipWithCompanyExecutives'] = tr.xpath('./td[3]/text()').get()
                manager_dict['listedCompany_companyEvent_executiveShareholdingchanges_variableQuantity'] = tr.xpath('./td[4]/span/text()').get().replace(" ", "")
                manager_dict['listedCompany_companyEvent_executiveShareholdingchanges_averageTransactionPrice'] = tr.xpath('./td[5]/text()').get()
                manager_dict['listedCompany_companyEvent_executiveShareholdingchanges_remainingShares'] = tr.xpath('./td[6]/text()').get()
                manager_dict['listedCompany_companyEvent_executiveShareholdingchanges_sharesChangeWays'] = tr.xpath('./td[7]/text()').get()
                # 加入列表
                item['listedCompany_companyEvent_executiveShareholdingchanges'].append(manager_dict)
        else:
            item['listedCompany_companyEvent_executiveShareholdingchanges'] = []



        # 3.股东持股变动模块
        # listedCompany_companyEvent_shareholderholdingchanges
        # 首先判断该公司有无该模块
        if response.xpath("//div[@id='holder']"):
            item['listedCompany_companyEvent_shareholderholdingchanges'] = list()
            tr_list = response.xpath("//div[@id='holder']//tbody/tr")
            for tr in tr_list:
                tr_dict = dict()
                tr_dict['listedCompany_companyEvent_shareholderholdingchanges_announcementDate'] = tr.xpath("./th/text()").get()
                if tr.xpath("./td[1]/span/text()").get():
                    tr_dict['listedCompany_companyEvent_shareholderholdingchanges_variableShareholder'] = tr.xpath("./td[1]/span/text()").get().strip()
                else:
                    tr_dict['listedCompany_companyEvent_shareholderholdingchanges_variableShareholder'] = tr.xpath("./td[1]/span/text()").get()
                tr_dict['listedCompany_companyEvent_shareholderholdingchanges_changeQuantity'] = tr.xpath("./td[2]/span/text()").get()
                if tr.xpath("./td[3]/text()"):
                    tr_dict['listedCompany_companyEvent_shareholderholdingchanges_averageTransactionPrice'] = tr.xpath("./td[3]/text()").get().strip()
                else:
                    tr_dict['listedCompany_companyEvent_shareholderholdingchanges_averageTransactionPrice'] = tr.xpath("./td[3]/text()").get()
                aaa = tr.xpath("./td[4]/text()").getall()
                tr_dict['listedCompany_companyEvent_shareholderholdingchanges_remainingSharesTotalNumber'] = "".join(aaa).strip()
                tr_dict['listedCompany_companyEvent_shareholderholdingchanges_changePeriod'] = tr.xpath("./td[5]/text()").get()
                tr_dict['listedCompany_companyEvent_shareholderholdingchanges_changeWays'] = tr.xpath("./td[6]/text()").get()
                # 加入列表
                item['listedCompany_companyEvent_shareholderholdingchanges'].append(tr_dict)
        else:
            item['listedCompany_companyEvent_shareholderholdingchanges'] = []



        yield item

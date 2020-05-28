# -*- coding: utf-8 -*-
# -*- coding: utf-8 -*-
import random
import re
from jqka_com_detail.jqka_com_detail.items import JqkaComDetailItem
from jqka_com_executive_introduction.items import JqkaComExecutiveIntroductionItem
from scrapy import item
from scrapy.http import Request
#获取excel需要引入的包
from openpyxl import load_workbook
from selenium import webdriver
import time
import scrapy


class JqkaComExecutiveIntroductSpider(scrapy.Spider):

    def __init__(self):
        self.driver = webdriver.PhantomJS(executable_path=r'C:\Users\10359\local\bin\phantomjs.exe')
        self.driver.set_page_load_timeout(40)

    name = 'jqka_com_executive_introduct'
    allowed_domains = ["basic.10.jqka.com"]

    def start_requests(self):
        yield Request("http://basic.10jqka.com.cn/603221/company.html", headers={
            'User-Agent': "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.87 Safari/537.36"})

    def parse(self, response):
        book = load_workbook(filename=r"C:\python\lnuSpider\data\exel\com_list.xlsx")
        sheet = book.active
        data = []
        data1 = []
        data2 = []
        row_num = 1
        while row_num <= 3:
            # 将表中第一列的1-100行数据写入data数组中
            data.append(sheet.cell(row=row_num, column=3).value)
            data1.append(sheet.cell(row=row_num, column=1).value)
            data2.append(row_num)
            row_num = row_num + 1
        for i in data2:
            # url = 'http://basic.10jqka.com.cn/'+data[i]+'/company.html'
            # print(data[i-1])
            listedCompany_url = 'http://basic.10jqka.com.cn/' + data[i - 1] + '/company.html'
            company_detail = JqkaComExecutiveIntroductionItem()
            company_detail['listedCompany_url'] = listedCompany_url
            listedCompany_id = data1[i - 1]
            company_detail['listedCompany_id'] = listedCompany_id
            # print(response.text)
            yield scrapy.Request(company_detail['listedCompany_url'],
                                 meta={'company_detail': company_detail}, callback=self.detail_ni, dont_filter=True)
        return

    def detail_ni(self, response):
        print("=====准备高管介绍中的信息====")
        #print(response.text)
        company_detail = response.meta['company_detail']
        root = response.xpath("//div[@class='content page_event_content']")[0]

        listedCompany_name = root.xpath("//div[@stat][@id='detail']//table[@class='m_table']//tr[1]""//td[2]//spa"
                                  "n/text()")[0].extract()
        company_detail['listedCompany_name'] = listedCompany_name

        print("=====高管信息准备完毕，休息一下嘻嘻====")
        root1 = response.xpath("//div[@id='manager'][@stat='company_manag"
                               "er']//div[@id='ml_001'][@class='m_tab_content']/"
                               "table[@class='m_table managelist m_hl']/tbody/tr[1]/td[@clas"
                               "s='tc name'][1]//table[@class='m_table ggintro']/thead/tr[1]//h3/text()").getall()
        # 这里要写个嵌套循环 首先对tr循环，其次对td进行循环 获取里面和外面的信息
        contents = response.xpath("//div[@id='manager'][@stat='company_manag"
                                  "er']//div[@id='ml_001'][@class='m_tab_content']/"
                                  "table[@class='m_table managelist m_hl']/tbody/tr")
        # print(type(contents))
        company_content1 = list()
        for content in contents:
            single = dict()
            content_1 = content.xpath("./td[1]/a/text()").getall()
            single['listedCompany_executiveInfor_name'] = "".join(content_1).strip()
            content_2 = content.xpath("./td[2]/text()").getall()
            single['listedCompany_executiveInfor_position'] = "".join(content_2).strip()
            content_3 = content.xpath("./td[3]//span/text()").getall()
            single['listedCompany_executiveInfor_direcShareNumbers'] = "".join(content_3).strip()
            content_4 = content.xpath("./td[4]//span/text()").getall()
            single['listedCompany_executiveInfor_indirecShareNumbers'] = "".join(content_4).strip()
            company_content1.append(single)
        company_detail['company_content1'] = company_content1

        company_content2 = list()
        for content in contents:
            single1 = dict()
            content_5 = content.xpath("./td[5]/a/text()").getall()
            single1['listedCompany_executiveInfor_name'] = "".join(content_5).strip()
            content_6 = content.xpath("./td[6]/text()").getall()
            single1['listedCompany_executiveInfor_position'] = "".join(content_6).strip()
            content_7 = content.xpath("./td[7]//span/text()").getall()
            single1['listedCompany_executiveInfor_direcShareNumbers'] = "".join(content_7).strip()
            content_8 = content.xpath("./td[8]//span/text()").getall()
            single1['listedCompany_executiveInfor_indirecShareNumbers'] = "".join(content_8).strip()
            company_content2.append(single1)
        company_detail['company_content2'] = company_content2

        contents1 = response.xpath("//div[@id='manager'][@stat='company_manag"
                                   "er']//div[@id='ml_002'][@class='m_tab_content']/"
                                   "table[@class='m_table managelist m_hl']/tbody/tr")
        company_content3 = list()
        for content in contents1:
            single2 = dict()
            content_1 = content.xpath("./td[1]/a/text()").getall()
            single2['listedCompany_executiveInfor_name'] = "".join(content_1).strip()
            content_2 = content.xpath("./td[2]/text()").getall()
            single2['listedCompany_executiveInfor_position'] = "".join(content_2).strip()
            content_3 = content.xpath("./td[3]//span/text()").getall()
            single2['listedCompany_executiveInfor_direcShareNumbers'] = "".join(content_3).strip()
            content_4 = content.xpath("./td[4]//span/text()").getall()
            single2['listedCompany_executiveInfor_indirecShareNumbers'] = "".join(content_4).strip()
            company_content3.append(single2)
        company_detail['company_content3'] = company_content3

        company_content4 = list()
        for content in contents1:
            single3 = dict()
            content_5 = content.xpath("./td[5]/a/text()").getall()
            single3['listedCompany_executiveInfor_name'] = "".join(content_5).strip()
            content_6 = content.xpath("./td[6]/text()").getall()
            single3['listedCompany_executiveInfor_position'] = "".join(content_6).strip()
            content_7 = content.xpath("./td[7]//span/text()").getall()
            single3['listedCompany_executiveInfor_direcShareNumbers'] = "".join(content_7).strip()
            content_8 = content.xpath("./td[8]//span/text()").getall()
            single3['listedCompany_executiveInfor_indirecShareNumbers'] = "".join(content_8).strip()
            company_content4.append(single3)
        company_detail['company_content4'] = company_content4

        contents3 = response.xpath("//div[@id='manager'][@stat='company_manag"
                                   "er']//div[@id='ml_003'][@class='m_tab_content']/"
                                   "table[@class='m_table managelist m_hl']/tbody/tr")
        company_content5 = list()
        for content in contents3:
            single4 = dict()
            content_1 = content.xpath("./td[1]/a/text()").getall()
            single4['listedCompany_executiveInfor_name'] = "".join(content_1).strip()
            content_2 = content.xpath("./td[2]/text()").getall()
            single4['listedCompany_executiveInfor_position'] = "".join(content_2).strip()
            content_3 = content.xpath("./td[3]//span/text()").getall()
            single4['listedCompany_executiveInfor_direcShareNumbers'] = "".join(content_3).strip()
            content_4 = content.xpath("./td[4]//span/text()").getall()
            single4['listedCompany_executiveInfor_indirecShareNumbers'] = "".join(content_4).strip()
            company_content5.append(single4)
        company_detail['company_content5'] = company_content5

        company_content6 = list()
        for content in contents3:
            single5 = dict()
            content_5 = content.xpath("./td[5]/a/text()").getall()
            single5['listedCompany_executiveInfor_name'] = "".join(content_5).strip()
            content_6 = content.xpath("./td[6]/text()").getall()
            single5['listedCompany_executiveInfor_position'] = "".join(content_6).strip()
            content_7 = content.xpath("./td[7]//span/text()").getall()
            single5['listedCompany_executiveInfor_direcShareNumbers'] = "".join(content_7).strip()
            content_8 = content.xpath("./td[8]//span/text()").getall()
            single5['listedCompany_executiveInfor_indirecShareNumbers'] = "".join(content_8).strip()
            company_content6.append(single5)
        company_detail['company_content6'] = company_content6

        time.sleep(random.randint(2, 5))
        yield company_detail


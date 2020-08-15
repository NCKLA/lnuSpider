# -*- coding: utf-8 -*-
import scrapy
from jqka_svzl.items import JqkaSvzlItem
from scrapy.http import Request
from openpyxl import load_workbook
import time
import random


class JqkaSvzlSpiderSpider(scrapy.Spider):
    name = 'jqka_svzl_spider'
    allowed_domains = ['http://basic.10jqka.com.cn/603290/position.html']

    def start_requests(self):
        yield Request("http://basic.10jqka.com.cn/603290/position.html", headers={
            'User-Agent': "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.87 Safari/537.36"})

    def parse(self, response):
        book = load_workbook(filename=r"C:\python\lnuSpider\data\exel\com_list.xlsx")
        sheet = book.active
        data = []
        data1 = []
        data2 = []
        data3 = []
        data4 = []
        row_num = 1
        while row_num <= 3815:
            # 将表中第一列的1-100行数据写入data数组中
            data.append(sheet.cell(row=row_num, column=3).value)
            data1.append(sheet.cell(row=row_num, column=1).value)
            data3.append(sheet.cell(row=row_num, column=2).value)
            data4.append(sheet.cell(row=row_num, column=4).value)
            data2.append(row_num)
            row_num = row_num + 1
        for i in data2:
            a = str(data[i - 1])
            listedCompany_url = 'http://basic.10jqka.com.cn/' + a + '/position.html'
            company_svzl= JqkaSvzlItem()
            company_svzl['listedCompany_url'] = listedCompany_url
            listedCompany_id = data1[i - 1]
            company_svzl['listedCompany_id'] = listedCompany_id
            listedCompany_name = data3[i - 1]
            company_svzl['listedCompany_name'] = listedCompany_name
            listedCompany_fullName = data4[i - 1]
            company_svzl['listedCompany_fullName'] = listedCompany_fullName
            yield scrapy.Request(company_svzl['listedCompany_url'],meta={'company_svzl': company_svzl}, callback=self.detail_ni, dont_filter=True)


    def detail_ni(self, response):
        # 获得相应公司对应的公司信息
        company_svzl = response.meta['company_svzl']
        item = JqkaSvzlItem()

        # 公司信息
        item['listedCompany_id'] = company_svzl['listedCompany_id']
        item['listedCompany_name'] = company_svzl['listedCompany_name']
        item['listedCompany_url'] = company_svzl['listedCompany_url']
        item['listedCompany_fullName'] = company_svzl['listedCompany_fullName']

        # 模块1： 上市公司_主力持仓_机构持股汇总
        item['listedCompany_svzl_institutionholdSummary'] = list()
        th_list = response.xpath("//div[@class='m_tab_content']//thead/tr/th")
        # 根据表达的数量进行遍历取值
        # 因为有一个是行头，所以需要-1
        for i in range(0, len(th_list)-1):
            td_dict = dict()
            td_dict['listedCompany_svzl_institutionholdSummary_reportingPeriod'] = response.xpath("//div[@class='m_tab_content']//thead/tr/th[{}]/text()".format(i+2)).get().strip()
            td_dict['listedCompany_svzl_institutionholdSummary_organizationNumber'] = response.xpath("//div[@class='m_tab_content']//tbody/tr[1]/td[{}]/text()".format(i+1)).get().strip()
            td_dict['listedCompany_svzl_institutionholdSummary_accumulatedHoldingQuantity'] = response.xpath("//div[@class='m_tab_content']//tbody/tr[2]/td[{}]/text()".format(i+1)).get().strip()
            td_dict['listedCompany_svzl_institutionholdSummary_totalMarketValue'] = response.xpath("//div[@class='m_tab_content']//tbody/tr[3]/td[{}]/text()".format(i+1)).get().strip()
            td_dict['listedCompany_svzl_institutionholdSummary_positionRatio'] = response.xpath("//div[@class='m_tab_content']//tbody/tr[4]/td[{}]/text()".format(i+1)).get().strip()
            # 较上期变化可能为空，空时源码发生变动，需要判断
            if response.xpath("//div[@class='m_tab_content']//tbody/tr[5]/td[{}]/span/text()".format(i+1)).get():
                td_dict['listedCompany_svzl_institutionholdSummary_comparedPreviousPeriodChange'] = response.xpath("//div[@class='m_tab_content']//tbody/tr[5]/td[{}]/span/text()".format(i+1)).get().strip()
            else:
                td_dict['listedCompany_svzl_institutionholdSummary_comparedPreviousPeriodChange'] = response.xpath("//div[@class='m_tab_content']//tbody/tr[5]/td[{}]/text()".format(i+1)).get().strip()
            # 把字典对象插入列表中
            item['listedCompany_svzl_institutionholdSummary'].append(td_dict)


        # 模块2： 上市公司_主力持仓_机构持股明细
        item['listedCompany_svzl_institutionholdDetail'] = list()
        # 日期数等于日期对应的div数目 len(holdetail_dates) = len(holdetail_divs)
        holdetail_dates = response.xpath("//div[@id='holdetail']//a[@class='fdate']/text()").getall()
        holdetail_divs = response.xpath("//div[@id='holdetail']//div[@class='m_tab_content clearfix pagination gssj_scroll_position']")
        for i in range(0, len(holdetail_divs)):
            # holdetail_div(每个日期)里面有N个invest_tr(1个invest_trs列表)
            holdetail_trs = holdetail_divs[i].xpath("./table[1]/tbody/tr")
            for holdetail_tr in holdetail_trs:
                # 通过字典来封装对象
                holdetail_dict = dict()
                # 日期 与 其他字段有所不同，直接从 invest_dates 列表 根据下标取就行
                holdetail_dict['listedCompany_svzl_institutionholdDetail_date'] = holdetail_dates[i]
                holdetail_dict['listedCompany_svzl_institutionholdDetail_organizationOrFundName'] = holdetail_tr.xpath("./th/span/text()").get().strip()
                holdetail_dict['listedCompany_svzl_institutionholdDetail_organizationType'] = holdetail_tr.xpath("./td[1]/text()").get().strip()
                holdetail_dict['listedCompany_svzl_institutionholdDetail_quantityHeld'] = holdetail_tr.xpath("./td[2]/text()").get().strip()
                holdetail_dict['listedCompany_svzl_institutionholdDetail_marketValue'] = holdetail_tr.xpath("./td[3]/text()").get().strip()
                holdetail_dict['listedCompany_svzl_institutionholdDetail_circulationSharesProportion'] = holdetail_tr.xpath("./td[4]/text()").get().strip()
                # 增减情况 不同情况源码不同，也需要判断
                 # 情况一：不发生增减
                if not holdetail_tr.xpath("./td[5]/span"):
                    holdetail_dict['listedCompany_svzl_institutionholdDetail_increaseOrDecrease'] = holdetail_tr.xpath("./td[5]/text()[2]").get().strip()
                # 情况二：发生增减
                else:
                    holdetail_dict['listedCompany_svzl_institutionholdDetail_increaseOrDecrease'] = holdetail_tr.xpath("./td[5]/span/text()").get().strip()
                # 基金收益排行 有的日期没有这个栏目，所以需要判断
                if holdetail_tr.xpath("./td[6]/text()"):
                    holdetail_dict['listedCompany_svzl_institutionholdDetail_fundIncomeRanking'] = holdetail_tr.xpath("./td[6]/text()").get().strip()
                else:
                    holdetail_dict['listedCompany_svzl_institutionholdDetail_fundIncomeRanking'] = '--'
                # 将封装好的对象添加到该模块的列表中
                item['listedCompany_svzl_institutionholdDetail'].append(holdetail_dict)


        # 模块3： 上市公司_主力持仓_IPO获配机构
        item['listedCompany_svzl_ipoInstitution'] = list()
        trs = response.xpath("//div[@class='bd pr']//tbody/tr")
        for tr in trs:
            tr_dict = dict()
            #tr_dict['listedCompany_svzl_ipoInstitution_id'] = tr.xpath("./th/text()").get()
            tr_dict['listedCompany_svzl_ipoInstitution_organizationName'] = tr.xpath("./td[@class='tl']/text()").get()
            tr_dict['listedCompany_svzl_ipoInstitution_allottedQuantity'] = tr.xpath("./td[2]/text()").get()
            tr_dict['listedCompany_svzl_ipoInstitution_applyPurchasingQuantity'] = tr.xpath("./td[3]/text()").get()
            tr_dict['listedCompany_svzl_ipoInstitution_lockupPeriod'] = tr.xpath("./td[@class='tc']/text()").get()
            tr_dict['listedCompany_svzl_ipoInstitution_organizationType'] = tr.xpath("./td[@class='tl']/text()").getall()[1]
            # 把字典对象插入列表中
            item['listedCompany_svzl_ipoInstitution'].append(tr_dict)

        yield item




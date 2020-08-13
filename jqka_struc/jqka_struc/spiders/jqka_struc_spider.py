# -*- coding: utf-8 -*-
import random
import re
from scrapy import item
from scrapy.http import Request
from openpyxl import load_workbook
from selenium import webdriver
import time
import scrapy
from jqka_struc.items import JqkaStrucItem


class JqkaStrucSpiderSpider(scrapy.Spider):
    name = 'jqka_struc_spider'
    allowed_domains = ['basic.10.jqka.com']
    start_urls = ['http://basic.10jqka.com.cn/603221/equity.html']

    def parse(self, response):
        # 读取excel表格
        book = load_workbook(filename=r"C:\python\lnuSpider\data\exel\com_list.xlsx")
        sheet = book.active
        data = []
        data1 = []
        data2 = []
        data3 = []
        data4 = []
        row_num = 1
        while row_num <= 3815:
            # 将表中第一列的1-100行数据写入data数组中
            data.append(sheet.cell(row=row_num, column=3).value)
            data1.append(sheet.cell(row=row_num, column=1).value)
            data3.append(sheet.cell(row=row_num, column=2).value)
            data4.append(sheet.cell(row=row_num, column=4).value)
            data2.append(row_num)
            row_num = row_num + 1
        for i in data2:
            a = str(data[i - 1])
            listedCompany_url = 'http://basic.10jqka.com.cn/' + a + '/equity.html'
            company_struc = JqkaStrucItem()
            company_struc['listedCompany_url'] = listedCompany_url
            listedCompany_id = data1[i - 1]
            company_struc['listedCompany_id'] = listedCompany_id
            listedCompany_name = data3[i - 1]
            company_struc['listedCompany_name'] = listedCompany_name
            listedCompany_fullName = data4[i - 1]
            company_struc['listedCompany_fullName'] = listedCompany_fullName
            yield scrapy.Request(company_struc['listedCompany_url'], meta={'company_struc': company_struc},callback=self.detail_ni, dont_filter=True)


    def detail_ni(self, response):
        # 获得相应公司对应的公司信息
        company_struc = response.meta['company_struc']
        item = JqkaStrucItem()

        # 公司信息
        item['listedCompany_id'] = company_struc['listedCompany_id']
        item['listedCompany_name'] = company_struc['listedCompany_name']
        item['listedCompany_url'] = company_struc['listedCompany_url']
        item['listedCompany_fullName'] = company_struc['listedCompany_fullName']

        # 1.解禁时间表 模块
        # listedCompany_struc_timeTable
        # 首先判断该公司有无该模块
        if response.xpath("//div[@id='liftban']"):
            item['listedCompany_equityStruc_liftBanTimeTable'] = list()
            tr_list = response.xpath("//div[@id='liftban']//tbody/tr")
            for tr in tr_list:
                tr_dict = dict()
                tr_dict['listedCompany_equityStruc_liftBanTimeTable_liftBanTime'] = tr.xpath("./th/text()").get()
                tr_dict['listedCompany_equityStruc_liftBanTimeTable_liftBanAnnouncedQuantity'] = tr.xpath("./td[1]/text()").get()
                tr_dict['listedCompany_equityStruc_liftBanTimeTable_liftBanSaleQuantity'] = tr.xpath("./td[2]/text()").get()
                tr_dict['listedCompany_equityStruc_liftBanTimeTable_liftBanSaleQuantityProportion'] = tr.xpath("./td[3]/text()").get()
                tr_dict['listedCompany_equityStruc_liftBanTimeTable_liftBanSharesTypes'] = tr.xpath("./td[4]/text()").get()
                tr_dict['listedCompany_equityStruc_liftBanTimeTable_previousDayClosingPrice'] = tr.xpath("./td[5]/text()").get()
                tr_dict['listedCompany_equityStruc_liftBanTimeTable_estimatedCost'] = tr.xpath("./td[6]/text()").get()
                tr_dict['listedCompany_equityStruc_liftBanTimeTable_announceValueOrNot'] = tr.xpath("./td[7]/text()").get()
                # 加入列表
                item['listedCompany_equityStruc_liftBanTimeTable'].append(tr_dict)
        else:
            item['listedCompany_equityStruc_liftBanTimeTable'] = []


        # 2.总股本结构 模块
        # listedCompany_equityStruc_shareStructure
        # 首先判断该公司有无该模块
        if response.xpath("//div[@id='stockcapit']"):
            item['listedCompany_equityStruc_shareStructure'] = list()
            date_list = response.xpath("//table[@class='mt15 m_table m_hl']/thead//th/text()").getall()
            tr_list = response.xpath("//table[@class='mt15 m_table m_hl']/tbody/tr")
            for i in range(1,len(date_list)):
                td_dict = dict()
                td_dict['listedCompany_equityStruc_shareStructure_time'] = date_list[i]
                # 初始化各个字段
                td_dict['listedCompany_equityStruc_shareStructure_totalShareCapital'] = ""
                td_dict['listedCompany_equityStruc_shareStructure_AShareCapital'] = ""
                td_dict['listedCompany_equityStruc_shareStructure_circulatingAShares'] = ""
                td_dict['listedCompany_equityStruc_shareStructure_restrictedAShares'] = ""
                td_dict['listedCompany_equityStruc_shareStructure_HShareCapital'] = ""
                td_dict['listedCompany_equityStruc_shareStructure_circulatingHShares'] = ""
                td_dict['listedCompany_equityStruc_shareStructure_restrictedHShares'] = ""
                td_dict['listedCompany_equityStruc_shareStructure_reasonForChange'] = ""
                for tr in tr_list:
                    # 该模块的字段除了日期都需要判断，因为有的公司字段不存在
                    if (tr.xpath("./th/text()").get()) == '总股本(股)':
                        td_dict['listedCompany_equityStruc_shareStructure_totalShareCapital'] = tr.xpath("./td[{}]/text()".format(i)).get()
                    if (tr.xpath("./th/text()").get()) == 'A股总股本(股)':
                        td_dict['listedCompany_equityStruc_shareStructure_AShareCapital'] = tr.xpath("./td[{}]/text()".format(i)).get()
                    if (tr.xpath("./th/span/text()").get()) == '流通A股(股)':
                        td_dict['listedCompany_equityStruc_shareStructure_circulatingAShares'] = tr.xpath("./td[{}]/text()".format(i)).get()
                    if (tr.xpath("./th/span/text()").get()) == '限售A股(股)':
                        td_dict['listedCompany_equityStruc_shareStructure_restrictedAShares'] = tr.xpath("./td[{}]/text()".format(i)).get()
                    if (tr.xpath("./th/text()").get()) == 'H股总股本(股)':
                        td_dict['listedCompany_equityStruc_shareStructure_HShareCapital'] = tr.xpath("./td[{}]/text()".format(i)).get()
                    if (tr.xpath("./th/text()").get()) == '   流通H股(股)':
                        td_dict['listedCompany_equityStruc_shareStructure_circulatingHShares'] = tr.xpath("./td[{}]/text()".format(i)).get()
                    if (tr.xpath("./th/text()").get()) == '   限售H股(股)':
                        td_dict['listedCompany_equityStruc_shareStructure_restrictedHShares'] = tr.xpath("./td[{}]/text()".format(i)).get()
                    if (tr.xpath("./th/text()").get()) == '   变动原因':
                        td_dict['listedCompany_equityStruc_shareStructure_reasonForChange'] = tr.xpath("./td[{}]/text()".format(i)).get()
                # 加入列表
                item['listedCompany_equityStruc_shareStructure'].append(td_dict)
        else:
            item['listedCompany_struc_totalEquityStructure'] = []


        # 3.股本变动 模块
        # ListedCompany_equityStruc_equityChanges
        # 首先判断该公司有无该模块
        if response.xpath("//div[@id='astockchange']"):
            item['ListedCompany_equityStruc_equityChanges'] = list()
            tr_list = response.xpath("//div[@id='astockchange']//tbody/tr")
            for tr in tr_list:
                tr_dict = dict()
                tr_dict['ListedCompany_equityStruc_equityChanges_date'] = tr.xpath("./td[1]/text()").get()
                tr_dict['ListedCompany_equityStruc_equityChanges_changeReason'] = tr.xpath("./td[2]/text()").get()
                tr_dict['ListedCompany_equityStruc_equityChanges_totalShareCapitalAShareAfterChange'] = tr.xpath("./td[3]/text()").get()
                tr_dict['ListedCompany_equityStruc_equityChanges_circulationASharesAfterChange'] = tr.xpath("./td[4]/text()").get()
                tr_dict['ListedCompany_equityStruc_equityChanges_restrictedASharesAfterChange'] = tr.xpath("./td[5]/text()").get()
                # 加入列表
                item['ListedCompany_equityStruc_equityChanges'].append(tr_dict)
        else:
            item['listedCompany_equityStruc_liftBanTimeTable'] = []



        yield item
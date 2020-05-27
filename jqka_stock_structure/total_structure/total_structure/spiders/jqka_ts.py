# -*- coding: utf-8 -*-
import random
import time

import scrapy
from openpyxl import load_workbook
from scrapy import Request

from jqka_stock_structure.total_structure.total_structure.items import TotalStructureItem


class JqkaTsSpider(scrapy.Spider):
    name = 'jqka_ts'
    allowed_domains = ['basic.10.jqka.com']

    def start_requests(self):
        yield Request("http://basic.10jqka.com.cn/603221/company.html", headers={
            'User-Agent': "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.87 Safari/537.36"})

    def parse(self, response):
        # 读取excel表格
        book = load_workbook(filename="com_list.xlsx")
        sheet = book.active
        data = []
        row_num = 1
        while row_num <= 500:

            data.append(sheet.cell(row=row_num, column=3).value)
            row_num = row_num + 1
        for i in data:
            jqka_ts1 = TotalStructureItem()
            url = 'http://basic.10jqka.com.cn/%s' % i + '/equity.html'
            jqka_ts1['url'] = url
            # print(response.text)
            yield scrapy.Request(jqka_ts1['url'], meta={'jqka_ts1': jqka_ts1}, callback=self.ts, dont_filter=True)
        return



    def ts(self, response):
        print("=====准备公司中的信息====")
        jqka_ts1 = response.meta['jqka_ts1']


        print("=====信息准备完毕====")

        contents = response.xpath("//div[@class='m_box gqtz'][@id='stockcapit']")
        content1 = contents.xpath("./div[@class='bd pt5']//table[@class='mt15 m_table m_hl']/tbody/tr")
        #contents = response.xpath("//div[@class='content page_event_content']/div[@class='m_box gqtz']")
        #content1 = contents.xpath("./div[@class='bd pt5']//table/tbody/tr")

        #print(content1)
        total_structure1_1 = list()
        for content in content1:
            single = dict()
            content_1 = content.xpath("./th[1]/text()").getall()
            single['content_1'] = "".join(content_1).strip()
            content_2 = content.xpath("./td[1]/text()").getall()
            single['content_2'] = "".join(content_2).strip()
            content_3 = content.xpath("./td[2]/text()").getall()
            single['content_3'] = "".join(content_3).strip()
            content_4 = content.xpath("./td[3]/text()").getall()
            single['content_4'] = "".join(content_4).strip()
            content_5 = content.xpath("./td[4]/text()").getall()
            single['content_5'] = "".join(content_5).strip()
            content_6 = content.xpath("./td[5]/text()").getall()
            single['content_6'] = "".join(content_6).strip()

            total_structure1_1.append(single)
        jqka_ts1['total_structure1_1'] = total_structure1_1
        #print(total_structure1_1)

        time.sleep(random.randint(2, 5))
        yield jqka_ts1

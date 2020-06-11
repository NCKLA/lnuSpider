# -*- coding: utf-8 -*-
import random
import time

import scrapy
from openpyxl import load_workbook
from scrapy import Request

from jqka_stock_structure.jqka_release_schedule.jqka_release_schedule.items import JqkaReleaseScheduleItem




class JqkaRsSpider(scrapy.Spider):
    name = 'jqka_rs'
    allowed_domains = ['basic.10.jqka.com']

    def start_requests(self):
        yield Request("http://basic.10jqka.com.cn/603221/equity.html", headers={
            'User-Agent': "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.87 Safari/537.36"})

    def parse(self, response):
        # 读取excel表格
        book = load_workbook(filename="com_list.xlsx")
        sheet = book.active
        data = []
        row_num = 1
        while row_num <= 500:
            data.append(sheet.cell(row=row_num, column=3).value)
            row_num = row_num + 1
        for i in data:
            jqka_rs1 = JqkaReleaseScheduleItem()
            url = 'http://basic.10jqka.com.cn/%s' % i + '/equity.html'
            jqka_rs1['url'] = url
            # print(response.text)
            yield scrapy.Request(jqka_rs1['url'],
                                 meta={'jqka_rs1': jqka_rs1}, callback=self.rs, dont_filter=True)
        return

    # options = webdriver.FirefoxOptions()  # 获取驱动配置信息

    # 修改/添加驱动信息
    # options.add_argument("--headless")  # 设置火狐为headless无界面模式
    # options.add_argument("--disable-gpu")  # 不使用gpu，我猜的，建议照抄这句

    # 如果驱动是放在根目录则不需要driver_path，同理下面的executable_path=driver_path也可以不加
    # driver_path = r"D:\python\geckodriver.exe"
    # driver = webdriver.Firefox(executable_path=driver_path, firefox_options=options)

    # 到这里，driver就是你的webdriver对象
    # driver.get('http://basic.10jqka.com.cn/603221/equity.html')

    def rs(self, response):
        print("=====准备公司中的信息====")
        jqka_rs1 = response.meta['jqka_rs1']

        print("=====信息准备完毕====")

        contents = response.xpath("//div[@class='content page_event_content']/div[@class='m_box main_intro']")
        content1 = contents.xpath("./div[@class='bd pt5']//table/tbody/tr")


        # print(response.text)
        release_schedule1_1 = list()
        for content in content1:
            single = dict()
            listedCompany_liftBanTimeTable_liftBanTime = content.xpath("./th[1]/text()").getall()
            single['listedCompany_liftBanTimeTable_liftBanTime'] = "".join(listedCompany_liftBanTimeTable_liftBanTime).strip()
            listedCompany_liftBanTimeTable_liftBanAnnouncedQuantity = content.xpath("./td[1]/text()").getall()
            single['listedCompany_liftBanTimeTable_liftBanAnnouncedQuantity'] = "".join(listedCompany_liftBanTimeTable_liftBanAnnouncedQuantity).strip()
            listedCompany_liftBanTimeTable_liftBanSaleQuantity = content.xpath("./td[2]/text()").getall()
            single['listedCompany_liftBanTimeTable_liftBanSaleQuantity'] = "".join(listedCompany_liftBanTimeTable_liftBanSaleQuantity).strip()
            listedCompany_liftBanTimeTable_liftBanSaleQuantityProportion = content.xpath("./td[3]/text()").getall()
            single['listedCompany_liftBanTimeTable_liftBanSaleQuantityProportion'] = "".join(listedCompany_liftBanTimeTable_liftBanSaleQuantityProportion).strip()
            listedCompany_liftBanTimeTable_liftBanSharesTypes = content.xpath("./td[4]/text()").getall()
            single['listedCompany_liftBanTimeTable_liftBanSharesTypes'] = "".join(listedCompany_liftBanTimeTable_liftBanSharesTypes).strip()
            listedCompany_liftBanTimeTable_previousDayClosingPrice = content.xpath("./td[5]/text()").getall()
            single['listedCompany_liftBanTimeTable_previousDayClosingPrice'] = "".join(listedCompany_liftBanTimeTable_previousDayClosingPrice).strip()
            listedCompany_liftBanTimeTable_estimatedCost = content.xpath("./td[6]/text()").getall()
            single['listedCompany_liftBanTimeTable_estimatedCost'] = "".join(listedCompany_liftBanTimeTable_estimatedCost).strip()
            listedCompany_liftBanTimeTable_announceValueOrNot = content.xpath("./td[7]/text()").getall()
            single['listedCompany_liftBanTimeTable_announceValueOrNot'] = "".join(listedCompany_liftBanTimeTable_announceValueOrNot).strip()

        release_schedule1_1.append(single)
        jqka_rs1['release_schedule1_1'] = release_schedule1_1

        time.sleep(random.randint(2, 5))
        yield jqka_rs1

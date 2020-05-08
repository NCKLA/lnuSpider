# -*- coding: utf-8 -*-
import random
import time

import scrapy
from jqka_news.items import JqkaNewsItem
from scrapy import item
from scrapy.http import Request
# 获取excel需要引入的包
from openpyxl import load_workbook
from selenium import webdriver
from selenium import webdriver
import scrapy
import json
import requests
import ip_proxy


class JqkaNewsSpiderSpider(scrapy.Spider):
    # def __init__(self):
    #     # self.driver = webdriver.PhantomJS(executable_path=r'C:\Users\10359\local\bin\phantomjs.exe')
    #     # self.driver.set_page_load_timeout(40)
    #     # self.driver = webdriver.PhantomJS(executable_path=r'C:\Users\G50\local\bin\phantomjs.exe')
    #     # # 设置timeout 不设置大概会像我一样卡死
    #     # self.driver.set_page_load_timeout(40)
    #     super().__init__()
    #     # self.min_page = 1
    #     # self.max_page = 100
    #     # self.list_page = range(self.min_page, self.max_page)
    #     self.json_obj = None
    #     self.domain = None
    #     self.port = None
    #     JqkaNewsSpiderSpider.new_port(self)

    # def new_port(self):
    #     print("准备开始获取url的try")
    #
    #     try:
    #         open_url = ip_proxy.get_open_url()
    #
    #         # 向代理服务器发起请求，去拿端口号（ip地址是固定的好像）
    #         r = requests.get(open_url, timeout=5)
    #         result = str(r.content)
    #         if "b\'" in result:
    #             result = result[2:-1]
    #         print("result   " + result)
    #         # logging.info('open_url||' + result)
    #         # json_obj为响应json
    #         self.json_obj = json.loads(result)
    #         code = self.json_obj['code']
    #         self.domain = self.json_obj['domain']
    #         # 获得的端口号（如果状态码为100）
    #         if code == 100:
    #             self.port = str(self.json_obj['port'][0])
    #         elif code == 108:
    #             reset_url = ip_proxy.get_reset_url()
    #             r = requests.get(reset_url, timeout=5)
    #         else:
    #             print("异常的状态码：" + str(code))
    #         # 状态码说明
    #         # 100 成功
    #         # 101 认证不通过
    #         # 102 请求格式不正确
    #         # 103 IP暂时耗尽
    #         # 106 账号使用时间到期
    #         # 118 ip使用量已用完
    #     except Exception as e:
    #         print("申请端口，try出事儿了" + repr(e))
    #     print("try完了")
    #     print("打印domain和port   " + self.domain + ":" + self.port)
    #
    # def close_port(self, port):
    #     print("准备开始关闭端口的try")
    #     try:
    #         print("开始try  准备close")
    #         close_url = ip_proxy.get_close_url(port)
    #         r = requests.get(close_url, timeout=5)
    #         print("close result: " + str(r.content))
    #     except Exception as e:
    #         print("关闭端口，try出事了: " + repr(e))

    name = 'jqka_news_spider'
    allowed_domains = ['basic.10jqka.com.cn']

    # start_urls = ['http://basic.10jqka.com.cn/603221/news.html']

    def start_requests(self):
        yield Request("http://basic.10jqka.com.cn/603221/news.html",
                      headers={
                          'User-Agent': "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.87 Safari/537.36"})

    # def start_requests(self):
    #     url = JqkaNewsSpiderSpider.start_urls[0]
    #     print(url)
    #     proxy = self.domain + ":" + str(self.port)
    #     proxies = ""
    #     if url.startswith("http://"):
    #         proxies = "http://" + str(proxy)
    #     elif url.startswith("https://"):
    #         proxies = "https://" + str(proxy)
    #     # 注意这里面的meta={'proxy':proxies},一定要是proxy进行携带,其它的不行,后面的proxies一定 要是字符串,其它任何形式都不行
    #     yield scrapy.Request(url, callback=self.parse, meta={'proxy': proxies})

    def parse(self, response):
        # 读取excel表格
        book = load_workbook(filename=r"C:\python\lnuSpider\data\exel\com_list.xlsx")
        sheet = book.active
        data = []
        row_num = 1
        while row_num <= 2:
            # 将表中第一列的1-100行数据写入data数组中
            data.append(sheet.cell(row=row_num, column=3).value)
            row_num = row_num + 1
        for i in data:
            company_news = JqkaNewsItem()
            com_url = 'http://basic.10jqka.com.cn/%s' % i + '/news.html'
            company_news['url'] = com_url
            # print(response.text)
            yield scrapy.Request(company_news['url'],
                                 meta={'company_news': company_news}, callback=self.detail_ni, dont_filter=True)
        return

    def detail_ni(self, response):
        company_news = response.meta['company_news']
        # root = response.xpath("//div[@class='content page_event_content']"
        #                       "/div[@class='m_box subnews'][@id='pub'][@stat='pub_pub']"
        #                       "/div[@class='bd']/div[@class='m_tab_content']/*").getall()
        # print(root)
        # root1 = response.xpath("//div[@class='m_dlbox'][@id='pull_all']/dl")
        # print(type(root1))
        # print(root1)

        options = webdriver.FirefoxOptions()
        options.add_argument("--headless")  # 设置火狐为headless无界面模式
        options.add_argument("--disable-gpu")

        driver_path = r"D:\project\geckodriver.exe"
        driver = webdriver.Firefox(executable_path=driver_path, firefox_options=options)
        driver.get(company_news['url'])

        while True:
            # c = driver.find_element_by_xpath("//div[@class='m_dlbox'][@id='pull_all']")
            # print(c.get_attribute('innerHTML'))
            root1 = driver.find_elements_by_xpath("//div[@class='m_dlbox'][@id='pull_all']//dl")

            # root1.get_attribute('innerHTML')
            # print(root1)
            for root2 in root1:
                news_date = root2.find_element_by_xpath("./dt/span[@class='date']")
                print(news_date.text)
                news_tag = root2.find_element_by_xpath("./dt/span[@class='title']/a/strong")
                print(news_tag.text)
                news_url = root2.find_element_by_xpath("./dt/span[@class='title']/a")
                print(news_url.get_attribute("href"))
                # print(news_tag)
                # print(news_url)
                # company_news['news_url'] = news_url.get_attribute("href")
                yield scrapy.Request(news_url.get_attribute("href"),
                                     meta={"news_url": news_url.get_attribute("href"), "url": company_news['url']},
                                     callback=self.detail_ni1,
                                     dont_filter=True)

            ul = driver.find_element_by_css_selector("[class='splpager clearfix light-theme simple-pagination']")
            # print(str(ul.tag_name))
            # div
            # print(ul.text)

            li_label_s = ul.find_elements_by_xpath('./ul/li')
            # print("测试 'innerHTML ："+li_label_s[-1].get_attribute('innerHTML'))
            xia_yi_ye = li_label_s[-1].get_attribute('innerHTML')
            if "下一页</a>" in xia_yi_ye:
                print("是a!")
                element = driver.find_element_by_css_selector("[class='page-link next']")
                element.click()
                time.sleep(3)
            elif "下一页</span>" in xia_yi_ye:
                print("是span!")
                break
            # try:
            #     c = driver.find_element_by_xpath("//div[@class='m_dlbox'][@id='pull_all']//a[@class='page-link next']")
            #     # print(c.get_attribute('innerHTML'))
            #     c.click()
            #     print("进入下一页")
            # except Exception as e:
            #     print("结束")
            #     break

            # if not c:
            #     print("===结束===")
            #     break
            # else:
            #     c.click()
            #     print("=====翻页成功======")

        # options = webdriver.FirefoxOptions()
        # options.set_headless(True)
        # options.add_argument("--headless")  # 设置火狐为headless无界面模式
        # options.add_argument("--disable-gpu")
        # driver = webdriver.Firefox(firefox_options=options)
        # driver.get('http://basic.10jqka.com.cn/603221/news.html')
        # < a href=" " class="page-link next">下一页</ a>
        # element = driver.find_element_by_css_selector("[class='page-link next']")
        # print(element.text)
        # element.click()
        # list1 = driver.find_element_by_id("pull_all").find_elements_by_class_name("client")
        # print(type(list1))
        # for one_news in list1:
        #     ss = one_news.get_attribute("href")
        #     print(ss)
        # print(news)
        # driver.close()
        # print(root3)

        # yield company_news

    # driver_path = r"C:\python\lnuSpider\msedgedriver.exe"
    # driver = webdriver.Ie(executable_path=driver_path)
    # driver.get('https://www.baidu.com/')
    # print(driver.page_source)

    def detail_ni1(self, response):
        print("======进入内页======")
        company_news = JqkaNewsItem()
        company_news['news_url'] = response.meta['news_url']
        company_news['url'] = response.meta['url']
        time.sleep(random.randint(1, 3))
        yield company_news

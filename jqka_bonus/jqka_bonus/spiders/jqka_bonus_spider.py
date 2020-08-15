# -*- coding: utf-8 -*-
import random
import re
from scrapy import item
from scrapy.http import Request
from openpyxl import load_workbook
from selenium import webdriver
import time
import scrapy
from jqka_bonus.items import JqkaBonusItem


class JqkaBonusSpiderSpider(scrapy.Spider):
    name = 'jqka_bonus_spider'
    allowed_domains = ['basic.10.jqka.com']
    start_urls = ['http://basic.10jqka.com.cn/603221/bonus.html']

    def parse(self, response):
        # 读取excel表格
        book = load_workbook(filename=r"C:\python\lnuSpider\data\exel\com_list.xlsx")
        sheet = book.active
        data = []
        data1 = []
        data2 = []
        data3 = []
        data4 = []
        row_num = 1
        while row_num <= 3815:
            # 将表中第一列的1-100行数据写入data数组中
            data.append(sheet.cell(row=row_num, column=3).value)
            data1.append(sheet.cell(row=row_num, column=1).value)
            data3.append(sheet.cell(row=row_num, column=2).value)
            data4.append(sheet.cell(row=row_num, column=4).value)
            data2.append(row_num)
            row_num = row_num + 1
        for i in data2:
            a = str(data[i - 1])
            listedCompany_url = 'http://basic.10jqka.com.cn/' + a + '/bonus.html'
            company_bonus = JqkaBonusItem()
            company_bonus['listedCompany_url'] = listedCompany_url
            listedCompany_id = data1[i - 1]
            company_bonus['listedCompany_id'] = listedCompany_id
            listedCompany_name = data3[i - 1]
            company_bonus['listedCompany_name'] = listedCompany_name
            listedCompany_fullName = data4[i - 1]
            company_bonus['listedCompany_fullName'] = listedCompany_fullName
            yield scrapy.Request(company_bonus['listedCompany_url'], meta={'company_bonus': company_bonus},callback=self.detail_ni, dont_filter=True)

    def detail_ni(self, response):
        # 获得相应公司对应的公司信息
        company_bonus = response.meta['company_bonus']
        item = JqkaBonusItem()

        # 公司信息
        item['listedCompany_id'] = company_bonus['listedCompany_id']
        item['listedCompany_name'] = company_bonus['listedCompany_name']
        item['listedCompany_url'] = company_bonus['listedCompany_url']
        item['listedCompany_fullName'] = company_bonus['listedCompany_fullName']

        # 1.分红情况 模块
        # listedCompany_dividendFinancing_dividendMatter
        # 首先判断该公司有无该模块
        if response.xpath("//div[@id='bonuslist']"):
            item['listedCompany_dividendFinancing_dividendMatter'] = list()
            tr_list = response.xpath("//div[@id='bonuslist']//tbody/tr")
            for tr in tr_list:
                tr_dict = dict()
                tr_dict['listedCompany_dividendFinancing_dividendMatter_generalMeetingDate'] = tr.xpath("./td[3]/text()").get()
                tr_dict['listedCompany_dividendFinancing_dividendMatter_implementationDate '] = tr.xpath("./td[4]/text()").get()
                tr_dict['listedCompany_dividendFinancing_dividendMatter_dividendPlanDescription '] = tr.xpath("./td[5]/text()").get()
                tr_dict['listedCompany_dividendFinancing_dividendMatter_AShareRegistrationDate'] = tr.xpath("./td[6]/text()").get()
                tr_dict['listedCompany_dividendFinancing_dividendMatter_AShareExDividendDate'] = tr.xpath("./td[7]/text()").get()
                tr_dict['listedCompany_dividendFinancing_dividendMatter_AShareDividendDate'] = tr.xpath("./td[8]/text()").get()
                tr_dict['listedCompany_dividendFinancing_dividendMatter_programmeProgress'] = tr.xpath("./td[9]/text()").get()
                tr_dict['listedCompany_dividendFinancing_dividendMatter_dividendPaymentRate'] = tr.xpath("./td[10]/text()").get()
                tr_dict['listedCompany_dividendFinancing_dividendMatter_dividendRate'] = tr.xpath("./td[11]/text()").get()
                # 加入列表
                item['listedCompany_dividendFinancing_dividendMatter'].append(tr_dict)
        else:
            item['listedCompany_dividendFinancing_dividendMatter'] = []


        # 2.配股情况 模块
        # listedCompany_dividendFinancing_allotmentCondition
        # 首先判断该公司有无该模块
        if response.xpath("//div[@id='stockallot']"):
            item['listedCompany_dividendFinancing_allotmentCondition'] = list()
            table_list= response.xpath("//div[@id='stockallot']//table[@class='m_table pggk mt10']")
            for table in table_list:
                table_dict = dict()
                table_dict['listedCompany_dividendFinancing_allotmentCondition_programmeProgress'] = table.xpath("./caption/span[1]/strong/text()").get()
                table_dict['listedCompany_companyEvent_executiveShareholdingchanges_variablePerson'] = table.xpath("./caption/span[2]/text()").get()
                table_dict['listedCompany_companyEvent_executiveShareholdingchanges_relationshipWithCompanyExecutives'] = table.xpath("./caption/span[3]/text()").get()
                table_dict['listedCompany_dividendFinancing_allotmentCondition_actualAllotmentProportion'] = table.xpath("./tbody/tr[1]/td[1]//strong/text()").get()
                table_dict['listedCompany_dividendFinancing_allotmentCondition_allotmentListingDate'] = table.xpath("./tbody/tr[1]/td[2]//strong/text()").get()
                table_dict['listedCompany_dividendFinancing_allotmentCondition_supervisionCommissionAnnouncementDate'] = table.xpath("./tbody/tr[1]/td[3]//strong/text()").get()
                table_dict['listedCompany_dividendFinancing_allotmentCondition_allotmentPricePerShare'] = table.xpath("./tbody/tr[2]/td[1]//strong/text()").get()
                table_dict['listedCompany_dividendFinancing_allotmentCondition_paymentStartAndEndDate'] = table.xpath("./tbody/tr[2]/td[2]/span/text()").get()
                table_dict['listedCompany_dividendFinancing_allotmentCondition_issuanceExaminationCommitteeAnnouncementDate'] = table.xpath("./tbody/tr[2]/td[3]/span/text()").get()
                table_dict['listedCompany_dividendFinancing_allotmentCondition_actuallyRaisedFundsNetAmount'] = table.xpath("./tbody/tr[3]/td[1]/span/text()").get()
                table_dict['listedCompany_dividendFinancing_allotmentCondition_plannedShareAllotmentProportionUpperLimit'] = table.xpath("./tbody/tr[4]/td[1]/span/text()").get()
                table_dict['listedCompany_dividendFinancing_allotmentCondition_ExDividendDate'] = table.xpath("./tbody/tr[4]/td[2]/span/text()").get()
                table_dict['listedCompany_dividendFinancing_allotmentCondition_shareholdersMeetingAnnouncementDate'] = table.xpath("./tbody/tr[4]/td[3]/span/text()").get()
                table_dict['listedCompany_dividendFinancing_allotmentCondition_fundRaisingAmountUpperLimit'] = table.xpath("./tbody/tr[5]/td[1]/span/text()").get()
                table_dict['listedCompany_dividendFinancing_allotmentCondition_equityRegistrationDate'] = table.xpath("./tbody/tr[5]/td[2]/span/text()").get()
                table_dict['listedCompany_dividendFinancing_allotmentCondition_directorsBoardAnnouncementDate'] = table.xpath("./tbody/tr[5]/td[3]/span/text()").get()
                # 加入列表
                item['listedCompany_dividendFinancing_allotmentCondition'].append(table_dict)
        else:
            item['listedCompany_dividendFinancing_allotmentCondition'] = []



        # 3.机构获配明细 模块
        # listedCompany_dividendFinancing_organizationAllocation
        # 首先判断该公司有无该模块
        if response.xpath("//div[@id='additionorgan']"):
            item['listedCompany_dividendFinancing_organizationAllocation'] = list()
            tr_list = response.xpath("//div[@id='additionorgan']//tbody/tr")
            for tr in tr_list:
                tr_dict = dict()
                tr_dict['listedCompany_dividendFinancing_organizationAllocation_organizationName'] = tr.xpath("./td[1]/text()").get()
                tr_dict['listedCompany_dividendFinancing_organizationAllocation_allottedQuantity'] = tr.xpath("./td[2]/text()").get()
                tr_dict['listedCompany_dividendFinancing_organizationAllocation_purchaseQuantity'] = tr.xpath("./td[3]/text()").get()
                tr_dict['listedCompany_dividendFinancing_organizationAllocation_lockupPeriod'] = tr.xpath("./td[4]/text()").get()
                tr_dict['listedCompany_dividendFinancing_organizationAllocation_liftingDate'] = tr.xpath("./td[5]/text()").get()
                tr_dict['listedCompany_dividendFinancing_organizationAllocation_organizationType'] = tr.xpath("./td[6]/text()").get()
                # 加入列表
                item['listedCompany_dividendFinancing_organizationAllocation'].append(tr_dict)
        else:
            item['listedCompany_dividendFinancing_organizationAllocation'] = []


        # 4.增发情况 模块
        # listedCompany_dividendFinancing_additionalIssue
        # 首先判断该公司有无该模块
        if response.xpath("//div[@id='additionprofile']"):
            item['listedCompany_dividendFinancing_additionalIssue'] = list()
            table_list = response.xpath("//div[@id='additionprofile']//table[@class='m_table pggk mt10']")
            for table in table_list:
                table_dict = dict()
                table_dict['listedCompany_dividendFinancing_additionalIssue_programmeProgress'] = table.xpath("./caption//strong/text()").get()
                table_dict['listedCompany_dividendFinancing_additionalIssue_issueCategory'] = table.xpath("./caption/span[2]/text()").get()
                table_dict['listedCompany_dividendFinancing_additionalIssue_issueMode'] = table.xpath("./caption/span[3]/text()").get()
                table_dict['listedCompany_dividendFinancing_additionalIssue_actualIssuePrice'] = table.xpath("./tbody/tr[1]/td[1]//strong/text()").get()
                table_dict['listedCompany_dividendFinancing_additionalIssue_listingAnnouncementDate'] = table.xpath("./tbody/tr[1]/td[2]//strong/text()").get()
                table_dict['listedCompany_dividendFinancing_additionalIssue_actualIssueNumber'] = table.xpath("./tbody/tr[2]/td[1]//strong/text()").get()
                table_dict['listedCompany_dividendFinancing_additionalIssue_newSharesIssueDate'] = table.xpath("./tbody/tr[2]/td[2]//strong/text()").get()
                table_dict['listedCompany_dividendFinancing_additionalIssue_actualNetFundraising'] = table.xpath("./tbody/tr[3]/td[1]/span/text()").get()
                table_dict['listedCompany_dividendFinancing_additionalIssue_supervisionCommissionAnnouncementDate'] = table.xpath("./tbody/tr[3]/td[2]/span/text()").get()
                table_dict['listedCompany_dividendFinancing_additionalIssue_planIssuePrice'] = table.xpath("./tbody/tr[4]/td[1]/span/text()").get()
                table_dict['listedCompany_dividendFinancing_additionalIssue_issuanceExaminationCommitteeAnnouncementDate'] = table.xpath("./tbody/tr[4]/td[2]/span/text()").get()
                table_dict['listedCompany_dividendFinancing_additionalIssue_issuePlansNumber'] = table.xpath("./tbody/tr[5]/td[1]/span/text()").get()
                table_dict['listedCompany_dividendFinancing_additionalIssue_shareholdersMeetingAnnouncementDate'] = table.xpath("./tbody/tr[5]/td[2]/span/text()").get()
                table_dict['listedCompany_dividendFinancing_additionalIssue_planFundRaisingAmount'] = table.xpath("./tbody/tr[6]/td[1]/span/text()").get()
                table_dict['listedCompany_dividendFinancing_additionalIssue_directorsBoardAnnouncementDate'] = table.xpath("./tbody/tr[6]/td[2]/span/text()").get()
                # 加入列表
                item['listedCompany_dividendFinancing_additionalIssue'].append(table_dict)
        else:
            item['listedCompany_dividendFinancing_additionalIssue'] = []


        yield item


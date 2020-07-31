# -*- coding: utf-8 -*-
import random
import re

from jqka_com_detail.items import JqkaComDetailItem
from scrapy import item
from scrapy.http import Request
# 获取excel需要引入的包
from openpyxl import load_workbook
from selenium import webdriver
import time
import scrapy


class JqkaComDetailSpiderSpider(scrapy.Spider):

    # def __init__(self):
    #     self.driver = webdriver.PhantomJS(executable_path=r'C:\Users\10359\local\bin\phantomjs.exe')
    #     self.driver.set_page_load_timeout(40)

    name = 'jqka_com_detail_spider'
    allowed_domains = ["basic.10.jqka.com"]
    start_urls = ["http://basic.10jqka.com.cn/603221/company.html"]

    # def start_requests(self):
    #     yield Request("http://basic.10jqka.com.cn/603221/company.html", headers={
    #         'User-Agent': "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.87 Safari/537.36"})

    def parse(self, response):
        # 读取excel表格
        book = load_workbook(filename=r"C:\python\lnuSpider\data\exel\com_list.xlsx")
        sheet = book.active
        data = []
        data1 = []
        data2 = []
        data3 = []
        row_num = 1
        while row_num <= 1000:
            # 将表中第一列的1-100行数据写入data数组中
            data.append(sheet.cell(row=row_num, column=3).value)
            data1.append(sheet.cell(row=row_num, column=1).value)
            data3.append(sheet.cell(row=row_num, column=2).value)
            data2.append(row_num)
            row_num = row_num + 1
        for i in data2:
            # url = 'http://basic.10jqka.com.cn/'+data[i]+'/company.html'
            # print(data[i-1])
            a=str(data[i - 1])
            print(type(a))
            listedCompany_url = 'http://basic.10jqka.com.cn/' + a + '/company.html'
            company_detail = JqkaComDetailItem()
            company_detail['listedCompany_url'] = listedCompany_url
            listedCompany_id = data1[i - 1]
            company_detail['listedCompany_id'] = listedCompany_id
            listedCompany_name = data3[i - 1]
            company_detail['listedCompany_name'] = listedCompany_name
            # print(listedCompany_id)
            yield scrapy.Request(company_detail['listedCompany_url'],
                                 meta={'company_detail': company_detail}, callback=self.detail_ni, dont_filter=True)
        return

    def detail_ni(self, response):
        print("=====准备详细情况中的第一部分信息====")
        company_detail = response.meta['company_detail']
        root = response.xpath("//div[@class='content page_event_content']")[0]
        # print(root)
        if root.xpath("//div[@stat][@id='detail']//table[@class='m_table']//tr[1]//td[2]//span/text()"):
            company_detail['listedCompany_enterpriseInfor_detailInfor_companyFullName'] = root.xpath("//div[@stat][@id='detail']//table[@class='m_table']//tr[1]//td[2]//span/text()")[0].extract()

            listedCompany_enterpriseInfor_detailInfor_region = root.xpath("//div[@stat][@id='detail']//table[@class='m_table']//tr[1]//td[3]//span/text()")[0].extract()
            company_detail['listedCompany_enterpriseInfor_detailInfor_region'] = listedCompany_enterpriseInfor_detailInfor_region
        else:
            company_detail['listedCompany_enterpriseInfor_detailInfor_region']='-'

        listedCompany_enterpriseInfor_detailInfor_englishName = root.xpath("//div[@stat][@id='detail']//table[@class='m_table']//tr[2]""//td[1]//span/text()")[0].extract()

        company_detail['listedCompany_enterpriseInfor_detailInfor_englishName'] = "".join(
            listedCompany_enterpriseInfor_detailInfor_englishName).strip().replace(' ', '')

        listedCompany_enterpriseInfor_detailInfor_industry = root.xpath("//div[@stat][@id='detail']//table[@class='m_table']//tr[2]//td[2]//span/text()")[0].extract()

        company_detail['listedCompany_enterpriseInfor_detailInfor_industry'] = listedCompany_enterpriseInfor_detailInfor_industry

        if root.xpath("//div[@stat][@id='detail']//table[@class='m_table']//tr[3]//td[1]//span/text()"):
            listedCompany_enterpriseInfor_detailInfor_formerlyUsedName = root.xpath("//div[@stat][@id='detail']//table[@class='m_table']//tr[3]//td[1]//span/text()")[0].extract()

            company_detail['listedCompany_enterpriseInfor_detailInfor_formerlyUsedName'] = listedCompany_enterpriseInfor_detailInfor_formerlyUsedName
        else:
            company_detail['listedCompany_enterpriseInfor_detailInfor_formerlyUsedName']='-'
        listedCompany_enterpriseInfor_detailInfor_websiteAddress = root.xpath("//div[@stat][@id='detail']//table[@class='m_table']//tr[3]//td[2]//span/a/text()")[0].extract()

        company_detail['listedCompany_enterpriseInfor_detailInfor_websiteAddress'] = listedCompany_enterpriseInfor_detailInfor_websiteAddress

        print("=====准备详细情况中的第二部分信息中的第一部分====")
        company_content2_1 = list()
        company_comments = response.xpath("//div[@stat][@id='detail']//div[@class='m_tab_content2']")[0]
        # print(response.text)
        # print(company_comments)
        single = dict()
        # 主营业务
        listedCompany_enterpriseInfor_detailInfor_mainBusiness = company_comments.xpath("//tr[1]/td[1]//span/text()").get()
        # print(company_main_business)
        single['listedCompany_enterpriseInfor_detailInfor_mainBusiness'] \
            = listedCompany_enterpriseInfor_detailInfor_mainBusiness.strip()
        # 产品名称
        listedCompany_enterpriseInfor_detailInfor_productName = company_comments.xpath("//tr[@class='product_name']//span/span/text()")[0].extract()

        listedCompany_enterpriseInfor_detailInfor_productName = "". \
            join(listedCompany_enterpriseInfor_detailInfor_productName).strip().replace(' ', '')
        # 去除字符串中存在的所有空格，调用re包
        if company_comments.xpath("//div[@class='tipbox_wrap mr10']//span/text()"):
            listedCompany_enterpriseInfor_detailInfor_productName = \
                re.sub('\s+', '', listedCompany_enterpriseInfor_detailInfor_productName).strip()
        # print(company_product_name)
            single['listedCompany_enterpriseInfor_detailInfor_productName'] = listedCompany_enterpriseInfor_detailInfor_productName.strip()
        else:
            single['listedCompany_enterpriseInfor_detailInfor_productName']='-'

        # 控股股东
        if company_comments.xpath("//div[@class='tipbox_wrap mr10']//span/text()"):
            listedCompany_enterpriseInfor_detailInfor_controllingShareholders = company_comments.xpath("//div[@class='tipbox_wrap mr10']//span/text()")[0].extract()

            single['listedCompany_enterpriseInfor_detailInfor_controllingShareholders'] = \
                "".join(listedCompany_enterpriseInfor_detailInfor_controllingShareholders).strip()
        else:
            single['listedCompany_enterpriseInfor_detailInfor_controllingShareholders']='-'
        # 实际控制人
        if company_comments.xpath("//tr[3]//span/text()"):
            listedCompany_enterpriseInfor_detailInfor_actualController = company_comments.xpath("//tr[3]//span/text()")[3].getall()
            single['listedCompany_enterpriseInfor_detailInfor_actualController'] = \
                "".join(listedCompany_enterpriseInfor_detailInfor_actualController).strip()
        else:
            single['listedCompany_enterpriseInfor_detailInfor_actualController']='-'

            # 最终控制人
        if company_comments.xpath("//tr[4]//span/text()"):
            listedCompany_enterpriseInfor_detailInfor_ultimateController = company_comments.xpath("//tr[4]//span/text()")[0].extract()
            single['listedCompany_enterpriseInfor_detailInfor_ultimateController'] = \
                "".join(listedCompany_enterpriseInfor_detailInfor_ultimateController).strip()
        else:
            single['listedCompany_enterpriseInfor_detailInfor_ultimateController']='-'

        company_content2_1.append(single)

        # print("子页面，打印这个字典=="+str(single))
        company_detail['company_content2_1'] = company_content2_1

        print("=====准备打印详细情况中的第二部分信息中的第二部分====")
        company_content2_2 = list()
        single1 = dict()
        if company_comments.xpath("//tr[5]//td[1]/span/a"):
            single1['listedCompany_enterpriseInfor_detailInfor_chairman'] = company_comments.xpath(
                "//tr[5]//td[1]/span/a/text()").get()
        else:
            single1['listedCompany_enterpriseInfor_detailInfor_chairman'] = '-'

        if company_comments.xpath("//tr[5]//td[2]/span/a"):
            single1['listedCompany_enterpriseInfor_detailInfor_chairmanSecretary'] = company_comments.xpath(
                "//tr[5]//td[2]/span/a/text()").get()
        else:
            single1['listedCompany_enterpriseInfor_detailInfor_chairmanSecretary'] = '-'

        if company_comments.xpath("//tr[5]//td[3]/span/a"):
            single1['listedCompany_enterpriseInfor_detailInfor_legalRepresentative'] = company_comments.xpath(
                "//tr[5]//td[3]/span/a/text()").get()
        else:
            single1['listedCompany_enterpriseInfor_detailInfor_legalRepresentative'] = '-'

        listedCompany_enterpriseInfor_detailInfor_generalManager = company_comments.xpath("//tr[6]//table[@class='m_table ggintro']//h3/text()").getall()
        single1['listedCompany_enterpriseInfor_detailInfor_generalManager'] = "".join(listedCompany_enterpriseInfor_detailInfor_generalManager).strip()

        listedCompany_enterpriseInfor_detailInfor_registeredCapital = company_comments.xpath("//tr[6]/td/span/text()")[3].getall()
        single1['listedCompany_enterpriseInfor_detailInfor_registeredCapital'] = "".join(listedCompany_enterpriseInfor_detailInfor_registeredCapital).strip()

        listedCompany_enterpriseInfor_detailInfor_staffNumbers = company_comments.xpath("//tr[6]/td/span/text()")[4].getall()
        single1['listedCompany_enterpriseInfor_detailInfor_staffNumbers'] = "".join(listedCompany_enterpriseInfor_detailInfor_staffNumbers).strip()

        listedCompany_enterpriseInfor_detailInfor_tel = company_comments.xpath("//tr[7]/td/span/text()")[0].getall()
        single1['listedCompany_enterpriseInfor_detailInfor_tel'] = "".join(listedCompany_enterpriseInfor_detailInfor_tel).strip()

        listedCompany_enterpriseInfor_detailInfor_fax = company_comments.xpath("//tr[7]/td/span/text()")[1].getall()
        single1['listedCompany_enterpriseInfor_detailInfor_fax'] = "".join(listedCompany_enterpriseInfor_detailInfor_fax).strip()

        listedCompany_enterpriseInfor_detailInfor_postalCode = company_comments.xpath("//tr[7]/td/span/text()")[2].getall()
        single1['listedCompany_enterpriseInfor_detailInfor_postalCode'] = "".join(listedCompany_enterpriseInfor_detailInfor_postalCode).strip()

        listedCompany_enterpriseInfor_detailInfor_businessAddress = company_comments.xpath("//tr[8]/td/span/text()").getall()
        single1['listedCompany_enterpriseInfor_detailInfor_businessAddress'] = "".join(listedCompany_enterpriseInfor_detailInfor_businessAddress).strip()
        company_content2_2.append(single1)
        company_detail['company_content2_2'] = company_content2_2

        print("=====准备详细情况中的第三部分信息====")
        listedCompany_enterpriseInfor_detailInfor_companyProfile = company_comments.xpath("//tr[9]//p[@class='tip lh24']/text()").getall()
        listedCompany_enterpriseInfor_detailInfor_companyProfile = "".join(listedCompany_enterpriseInfor_detailInfor_companyProfile).strip()
        company_detail['listedCompany_enterpriseInfor_detailInfor_companyProfile'] = listedCompany_enterpriseInfor_detailInfor_companyProfile

        print("=====详细情况准备完毕，休息一下嘻嘻====")
        time.sleep(random.randint(1, 3))
        yield company_detail

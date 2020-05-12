# -*- coding: utf-8 -*-
import scrapy
import random
import re
from jqka_lssuance_related.items import JqkaLssuanceRelatedItem
from scrapy import item
from scrapy.http import Request
# 获取excel需要引入的包
from openpyxl import load_workbook
from selenium import webdriver
import time
import scrapy


class JqkaLrSpider(scrapy.Spider):
    name = 'jqka_lr'

    # allowed_domains = ['http://basic.10jqka.com.cn/603221/company.html']
    # start_urls = ['http://http://basic.10jqka.com.cn/603221/company.html/']
    def __init__(self):
        self.driver = webdriver.PhantomJS(executable_path=r'C:\Users\10359\local\bin\phantomjs.exe')
        self.driver.set_page_load_timeout(40)

    allowed_domains = ["basic.10.jqka.com"]

    def start_requests(self):
        yield Request("http://basic.10jqka.com.cn/603221/company.html", headers={
            'User-Agent': "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.87 Safari/537.36"})

    def parse(self, response):
        # 读取excel表格
        book = load_workbook(filename=r"C:\python\lnuSpider\data\exel\com_list.xlsx")
        sheet = book.active
        data = []
        data1 = []
        data2 = []
        data3 = []
        row_num = 1
        while row_num <= 3:
            # 将表中第一列的1-100行数据写入data数组中
            data.append(sheet.cell(row=row_num, column=3).value)
            data1.append(sheet.cell(row=row_num, column=1).value)
            data3.append(sheet.cell(row=row_num, column=2).value)
            data2.append(row_num)
            row_num = row_num + 1

            # print(response.text)
        for i in data2:
            # url = 'http://basic.10jqka.com.cn/'+data[i]+'/company.html'
            # print(data[i-1])
            listedCompany_url = 'http://basic.10jqka.com.cn/' + data[i - 1] + '/company.html'
            company_lr = JqkaLssuanceRelatedItem()
            company_lr['listedCompany_url'] = listedCompany_url
            listedCompany_id = data1[i - 1]
            company_lr['listedCompany_id'] = listedCompany_id
            listedCompany_name = data3[i - 1]
            company_lr['listedCompany_name'] = listedCompany_name
            # print(response.text)
            #
            yield scrapy.Request(company_lr['listedCompany_url'],
                                 meta={'company_lr': company_lr}, callback=self.detail_ni, dont_filter=True)
        return

    def detail_ni(self, response):
        print("=====准备详细情况中的第一部分信息====")
        company_lr = response.meta['company_lr']
        root = response.xpath("//div[@class='content page_event_content']")[0]
        # print(root)
        # company_lr['name'] = root.xpath("//div[@stat][@id='detail']//table[@class='m_table']//tr[1]""//td[2]//spa"
        #                                     "n/text()")[0].extract()
        content1 = response.xpath("//div[@id='publish'][@stat='company_publish']/div[@class='bd pr']/"
                                  "table[@class='m_table']//tr[1]")

        listedCompany_issueRelated_establishmentDate = content1.xpath("./td[1]/span/text()").getall()
        company_lr['listedCompany_issueRelated_establishmentDate'] = listedCompany_issueRelated_establishmentDate

        listedCompany_issueRelated_issueNumber = content1.xpath("./td[2]/span/text()").getall()
        company_lr['listedCompany_issueRelated_issueNumber'] = listedCompany_issueRelated_issueNumber

        listedCompany_issueRelated_issuePrice = content1.xpath("./td[3]/span/text()").getall()
        company_lr['listedCompany_issueRelated_issuePrice'] = listedCompany_issueRelated_issuePrice

        content2 = response.xpath("//div[@id='publish'][@stat='company_publish']/div[@class='bd pr']/"
                                  "table[@class='m_table']//tr[2]")
        listedCompany_issueRelated_listingDate = content2.xpath("./td[1]/span/text()").getall()
        company_lr['listedCompany_issueRelated_listingDate'] = listedCompany_issueRelated_listingDate

        listedCompany_issueRelated_issuePriceEarningsRatio = content2.xpath("./td[2]/span/text()").getall()
        company_lr['listedCompany_issueRelated_issuePriceEarningsRatio'] = listedCompany_issueRelated_issuePriceEarningsRatio

        listedCompany_issueRelated_expectedFundraising = content2.xpath("./td[3]/span/text()").getall()
        company_lr['listedCompany_issueRelated_expectedFundraising'] = listedCompany_issueRelated_expectedFundraising

        content3 = response.xpath("//div[@id='publish'][@stat='company_publish']/div[@class='bd pr']/"
                                  "table[@class='m_table']//tr[3]")

        listedCompany_issueRelated_firstDayOpeningPrice = content3.xpath("./td[1]/span/text()").getall()
        company_lr['listedCompany_issueRelated_firstDayOpeningPrice'] = listedCompany_issueRelated_firstDayOpeningPrice

        listedCompany_issueRelated_IssuanceRate = content3.xpath("./td[2]/span/text()").getall()
        company_lr['listedCompany_issueRelated_IssuanceRate'] = listedCompany_issueRelated_IssuanceRate

        listedCompany_issueRelated_actualFundraising = content3.xpath("./td[3]/span/text()").getall()
        company_lr['listedCompany_issueRelated_actualFundraising'] = listedCompany_issueRelated_actualFundraising

        content4 = response.xpath("//div[@id='publish'][@stat='company_publish']/div[@class='bd pr']/"
                                  "table[@class='m_table']//tr[4]")

        listedCompany_issueRelated_leadUnderwriter = content4.xpath("//div[@class='main_sell'][1]/span/text()").getall()
        company_lr['listedCompany_issueRelated_leadUnderwriter'] = listedCompany_issueRelated_leadUnderwriter

        listedCompany_issueRelated_listingSponsor = content4.xpath("//div[@class='main_sell'][2]/span/text()").getall()
        company_lr['listedCompany_issueRelated_listingSponsor'] = listedCompany_issueRelated_listingSponsor

        listedCompany_issueRelated_history = response.xpath("//div[@id='publish'][@stat='company_publish']//p[@class='tip lh24']/text()").getall()
        listedCompany_issueRelated_history = "".join(listedCompany_issueRelated_history).strip().replace(' ', '')
        company_lr['listedCompany_issueRelated_history'] = listedCompany_issueRelated_history

        time.sleep(random.randint(2, 5))
        yield company_lr

# -*- coding: utf-8 -*-
import scrapy
from jqka_Ri_event.items import JqkaRiEventItem
# -*- coding: utf-8 -*-
import scrapy
import random
import re
from scrapy import item
from scrapy.http import Request
# 获取excel需要引入的包
from openpyxl import load_workbook
from selenium import webdriver
import time
import scrapy


class JqkaRieSpider(scrapy.Spider):
    def __init__(self):
        self.driver = webdriver.PhantomJS(executable_path=r'C:\Users\10359\local\bin\phantomjs.exe')
        self.driver.set_page_load_timeout(40)

    name = 'jqka_rie'
    allowed_domains = ["basic.10.jqka.com"]

    def start_requests(self):
        yield Request("http://basic.10jqka.com.cn/603221/index.html", headers={
            'User-Agent': "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.87 Safari/537.36"})

    def parse(self, response):
        # 读取excel表格
        book = load_workbook(filename=r"C:\python\lnuSpider\data\exel\com_list.xlsx")
        sheet = book.active
        data = []
        row_num = 1
        while row_num <= 1:
            # 将表中第一列的1-100行数据写入data数组中
            data.append(sheet.cell(row=row_num, column=3).value)
            row_num = row_num + 1
        for i in data:
            company_co = JqkaRiEventItem()
            url = 'http://basic.10jqka.com.cn/%s' % i + '/index.html'
            company_co['url'] = url
            # print(response.text)
            yield scrapy.Request(company_co['url'],
                                 meta={'company_co': company_co}, callback=self.detail_ni, dont_filter=True)
        return

    def detail_ni(self, response):
        company_1 = list()
        single = dict()
        company_co = response.meta['company_co']
        content1 = response.xpath("//div[@class='m_box event new_msg z101']"
                                  "[@id='pointnew']//table"
                                  "[@class='m_table m_table_db'][@id='tableToday']/tbody/tr[1]")
        content11 = "".join(content1.xpath("./td[2]/span/a/text()").getall()).strip()
        # print(content11)
        single['content11'] = content11
        content2 = response.xpath("//div[@class='m_box event new_msg z101']"
                                  "[@id='pointnew']//table"
                                  "[@class='m_table m_table_db'][@id='tableToday']/tbody/tr[2]")
        content12 = "".join(content2.xpath("./td/span/a/text()").getall()).strip()
        single['content12'] = content12
        # print(content12)

        content3 = response.xpath("//div[@class='m_box event new_msg z101']"
                                  "[@id='pointnew']//table"
                                  "[@class='m_table m_table_db'][@id='tableToday']/tbody/tr[3]")
        content13 = "".join(content3.xpath("./td/span/a/text()").getall()).strip()
        single['content13'] = content13
        company_1.append(single)
        # print(content13)
        company_co['company_1'] = company_1
        content4 = response.xpath("//div[@class='m_box event new_msg z101']"
                                  "[@id='pointnew']//table"
                                  "[@class='m_table m_table_db'][@id='tableList']/tbody/tr")
        for content_2 in content4:
            content_22 = content_2.xpath("./td[2]/span/a/text()").getall()
            content_22 = "".join(content_22).strip().replace(' ', '')
            print(content_22)
        yield company_co

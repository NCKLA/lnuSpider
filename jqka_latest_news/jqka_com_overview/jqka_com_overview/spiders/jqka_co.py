# -*- coding: utf-8 -*-
import scrapy
import random
import re

from jqka_com_overview.items import JqkaComOverviewItem
from pandas._libs import json
from scrapy import item
from scrapy.http import Request
# 获取excel需要引入的包
from openpyxl import load_workbook
from selenium import webdriver
import time
import scrapy


class JqkaCoSpider(scrapy.Spider):

    name = 'jqka_co'
    allowed_domains = ["basic.10.jqka.com"]

    def start_requests(self):
        yield Request("http://basic.10jqka.com.cn/603221/index.html", headers={
            'User-Agent': "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.87 Safari/537.36"})

    def parse(self, response):
        # 读取excel表格
        book = load_workbook(filename=r"C:\python\lnuSpider\data\exel\com_list.xlsx")
        sheet = book.active
        data = []
        data1 = []
        data2 = []
        data3 = []
        row_num = 1
        while row_num <= 3815:
            # 将表中第一列的1-100行数据写入data数组中
            data.append(sheet.cell(row=row_num, column=3).value)
            data1.append(sheet.cell(row=row_num, column=1).value)
            data3.append(sheet.cell(row=row_num, column=2).value)
            data2.append(row_num)
            row_num = row_num + 1
        for i in data2:
            # url = 'http://basic.10jqka.com.cn/'+data[i]+'/company.html'
            # print(data[i-1])
            a = str(data[i - 1])
            listedCompany_url = 'http://basic.10jqka.com.cn/' + a + '/index.html'
            url = 'http://basic.10jqka.com.cn/mapp/' + a + '/a_stock_foucs.json'
            url1 = 'http://basic.10jqka.com.cn/mapp/' + a + '/a_companies_list.json'
            company_co = JqkaComOverviewItem()
            company_co['listedCompany_url'] = listedCompany_url
            listedCompany_id = data1[i - 1]
            company_co['listedCompany_id'] = listedCompany_id
            listedCompany_name = data3[i - 1]
            company_co['listedCompany_name'] = listedCompany_name
            # print(listedCompany_id)
            yield scrapy.Request(company_co['listedCompany_url'],
                                 meta={'company_co': company_co, "url": url, "url1": url1}, callback=self.detail_ni, dont_filter=True)
        return

    def detail_ni(self, response):
        company_co = response.meta['company_co']
        url = response.meta['url']
        url1 = response.meta['url1']
        # #============公司概要==============
        root = response.xpath("//div[@class='content page_event_content']/div[@id='profile']/div[@class='bd']/table[1]")
        root1 = root.xpath("./tbody/tr[1]/td[1]/span[2]/text()").getall()
        listedCompany_latestNews_comBrief_comHighlights = ''.join(root1).strip()
        company_co['listedCompany_latestNews_comBrief_comHighlights'] = listedCompany_latestNews_comBrief_comHighlights
        #
        # root2 = root.xpath("./tbody/tr[1]/td[2]/span[2]/text()").getall()
        # listedCompany_latestNews_comBrief_comPopularityRanking = ''.join(root2).strip()
        # company_co['listedCompany_latestNews_comBrief_comPopularityRanking'] = listedCompany_latestNews_comBrief_comPopularityRanking
        #
        # root3 = root.xpath("./tbody/tr[1]/td[2]/span[4]/text()").getall()
        # listedCompany_latestNews_comBrief_industryPopularityRanking = ''.join(root3).strip()
        # company_co['listedCompany_latestNews_comBrief_industryPopularityRanking'] = listedCompany_latestNews_comBrief_industryPopularityRanking
        #
        root4 = root.xpath("./tbody/tr[2]/td[1]/span[2]/a/text()").getall()
        listedCompany_latestNews_comBrief_mainBusiness = ''.join(root4).strip()
        company_co['listedCompany_latestNews_comBrief_mainBusiness'] = listedCompany_latestNews_comBrief_mainBusiness
        #
        root5 = root.xpath("./tbody/tr[2]/td[2]/span[2]/text()").getall()
        listedCompany_latestNews_comBrief_shenwanIndustry = ''.join(root5).strip()
        company_co['listedCompany_latestNews_comBrief_shenwanIndustry'] = listedCompany_latestNews_comBrief_shenwanIndustry

        root6 = root.xpath("./tbody/tr[3]/td/div[2]/a/text()").getall()
        listedCompany_latestNews_comBrief_conceptFitRanking = ''.join(root6).strip()
        company_co['listedCompany_latestNews_comBrief_conceptFitRanking'] = listedCompany_latestNews_comBrief_conceptFitRanking

        # root7 = root.xpath("./tbody/tr[4]/td/div[3]/div[1]/div[1]/a/text()").getall()
        # listedCompany_latestNews_comBrief_domesticComparableCompanies = ''.join(root7).strip()
        # company_co['listedCompany_latestNews_comBrief_domesticComparableCompanies'] = listedCompany_latestNews_comBrief_domesticComparableCompanies
        # #
        # root2 = root.xpath("./tbody/tr[2]/td[1]/span[2]/a/text()").getall()
        # print(root2)
        # root3 = root.xpath("./tbody/tr[2]/td[2]/span[2]/text()").getall()
        # print(root3)
        #
        # root3 = root.xpath("./tbody/tr[3]/td/div[2]/a")
        # list1 = list()
        # for rot in root3:
        #     cont = rot.xpath("./text()").getall()
        #     # print(cont)
        #     list1.append(cont)
        # print(list1)
        #
        # root = response.xpath("//div[@class='content page_event_content']/div[@id='profile']/div[@class='bd']/table[2]")
        # root1 = root.xpath("./tbody/tr[1]/td[1]/span[2]/text()").getall()
        # print(root1)
        # root1 = root.xpath("./tbody/tr[1]/td[2]/span[2]/text()").getall()
        # print(root1)
        # root1 = root.xpath("./tbody/tr[1]/td[3]/span[2]/text()").getall()
        # print(root1)
        # root1 = root.xpath("./tbody/tr[1]/td[4]/span[2]/text()").getall()
        # print(root1)
        # root1 = root.xpath("./tbody/tr[2]/td[1]/span[2]/text()").getall()
        # print(root1)
        # root1 = root.xpath("./tbody/tr[2]/td[2]/span[2]/text()").getall()
        # print(root1)
        # root1 = root.xpath("./tbody/tr[2]/td[3]/span[2]/text()").getall()
        # print(root1)
        # root1 = root.xpath("./tbody/tr[2]/td[4]/span[2]/text()").getall()
        # print(root1)
        yield scrapy.Request(url=url, meta={'company_co': company_co, "url1": url1}, callback=self.detail_parse, dont_filter=True)

    def detail_parse(self, response):
        company_co = response.meta['company_co']
        url1 = response.meta['url1']
        art1 = response.text
        art = json.loads(art1)
        # 转换成字典形式
        company_co['listedCompany_latestNews_comBrief_comPopularityRanking'] = art['data']['all_rank']
        # print(art)
        # print(type(art))
        # print(art['data']['all_rank'])
        company_co['listedCompany_latestNews_comBrief_industryPopularityRanking'] = art['data']['industry_rank']
        # 获取其中的标签
        # print(art['data']['industry_rank'])
        # all_rank=art['data'][0]['all_rank']
        # print(all_rank)
        yield scrapy.Request(url=url1, meta={'company_co': company_co}, callback=self.detail_parse1, dont_filter=True)

    def detail_parse1(self, response):
        company_co = response.meta['company_co']
        art2 = response.text
        art = json.loads(art2)
        print(art)
        a = art['data']['domestic']['company_data'][0]['list']
        # print(a)
        index = 0
        list1=list()
        while index < len(a):
            # print(a[index]['name'])
            list1.append(a[index]['name'])
            index += 1
        # print(list1)
        a = dict(list(enumerate(list1)))
        company_co['listedCompany_latestNews_comBrief_domesticMarketsComparableCompanies'] = a
        # 截取之后的

        a1 = art['data']['abroad']['company_data'][0]['list']
        index = 0
        list2 = list()
        while index < len(a1):
            # print(a[index]['name'])
            a1[index]['name'] = ''.join(a1[index]['name']).strip().replace(' ', '')
            list2.append(a1[index]['name'])
            index += 1
        a1 = dict(list(enumerate(list2)))
        company_co['listedCompany_latestNews_comBrief_foreignMarketsComparableCompanies'] = a1

        time.sleep(random.randint(1, 3))
        yield company_co


# -*- coding: utf-8 -*-
import scrapy
# -*- coding: utf-8 -*-
import scrapy
import random
import re

from jqka_news_announcement.items import JqkaNewsAnnouncementItem
from scrapy import item
from scrapy.http import Request
# 获取excel需要引入的包
from openpyxl import load_workbook
from selenium import webdriver
import time
import scrapy


class JqkaNaSpider(scrapy.Spider):
    def __init__(self):
        self.driver = webdriver.PhantomJS(executable_path=r'C:\Users\10359\local\bin\phantomjs.exe')
        self.driver.set_page_load_timeout(40)

    name = 'jqka_na'
    allowed_domains = ["basic.10.jqka.com"]

    def start_requests(self):
        yield Request("http://basic.10jqka.com.cn/603221/index.html", headers={
            'User-Agent': "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.87 Safari/537.36"})

    def parse(self, response):
        # 读取excel表格
        book = load_workbook(filename=r"C:\python\lnuSpider\data\exel\com_list.xlsx")
        sheet = book.active
        data = []
        row_num = 1
        while row_num <= 1:
            # 将表中第一列的1-100行数据写入data数组中
            data.append(sheet.cell(row=row_num, column=3).value)
            row_num = row_num + 1
        for i in data:
            company_na = JqkaNewsAnnouncementItem()
            url = 'http://basic.10jqka.com.cn/%s' % i + '/index.html'
            company_na['url'] = url
            # print(response.text)
            yield scrapy.Request(company_na['url'],
                                 meta={'company_na': company_na}, callback=self.detail_ni, dont_filter=True)
        return

    def detail_ni(self, response):
        company_na = response.meta['company_na']

        content1 = response.xpath("//div[@class='m_box post_news fl post clearfix']//ul/li")
        # print(content1)
        for content_1 in content1:
            content_11 = content_1.xpath("./a/text()").getall()
            content_11 = "".join(content_11).strip()
            content_12 = content_1.xpath("./a/span/text()").getall()
            content_12 = "".join(content_12).strip()
            print(content_11)
            print(content_12)

        content2 = response.xpath("//div[@class='m_box comp_post fr post']//ul/li")
        # print(content1)
        for content_2 in content2:
            content_21 = content_2.xpath("./a/text()").getall()
            content_21 = "".join(content_21).strip()
            content_22 = content_2.xpath("./a/span/text()").getall()
            content_22 = "".join(content_22).strip()
            print(content_21)
            print(content_22)

        yield company_na


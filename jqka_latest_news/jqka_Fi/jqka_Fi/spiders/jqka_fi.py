# -*- coding: utf-8 -*-
import random
import re
from scrapy import item
from scrapy.http import Request
#获取excel需要引入的包
from openpyxl import load_workbook
from selenium import webdriver
import time
import scrapy
from jqka_Fi.items import JqkaFiItem



class JqkaFiSpider(scrapy.Spider):
    name = 'jqka_fi'
    allowed_domains = ['basic.10.jqka.com']

    def start_requests(self):
        yield Request("http://basic.10jqka.com.cn/603221/index.html", headers={
            'User-Agent': "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.87 Safari/537.36"})

    def parse(self, response):
        # 读取excel表格
        book = load_workbook(filename="com_list.xlsx")
        sheet = book.active
        data = []
        row_num = 1
        while row_num <= 500:

            data.append(sheet.cell(row=row_num, column=3).value)
            row_num = row_num + 1
        for i in data:
            jqka_fi1 = JqkaFiItem()
            url = 'http://basic.10jqka.com.cn/%s' % i + '/index.html'
            jqka_fi1['url'] = url
            # print(response.text)
            yield scrapy.Request(jqka_fi1['url'],
                                 meta={'jqka_fi1': jqka_fi1}, callback=self.fi, dont_filter=True)
        return



    def fi(self, response):
        print("=====准备公司中的信息====")
        jqka_fi1 = response.meta['jqka_fi1']


        print("=====信息准备完毕====")

        contents = response.xpath("//div[@class='m_box new_msg z100'][@id='finance']"
                                  "//table[@class='m_table m_hl fixtable']/tbody/tr")

        finance_index1_1 = list()
        for content in contents:
            single = dict()
            content_1 = content.xpath("./td[1]/text()").getall()
            single['content_1'] = "".join(content_1).strip()
            content_2 = content.xpath("./td[2]/text()").getall()
            single['content_2'] = "".join(content_2).strip()
            content_3 = content.xpath("./td[3]/text()").getall()
            single['content_3'] = "".join(content_3).strip()
            content_4 = content.xpath("./td[4]/text()").getall()
            single['content_4'] = "".join(content_4).strip()
            content_5 = content.xpath("./td[5]/text()").getall()
            single['content_5'] = "".join(content_5).strip()
            content_6 = content.xpath("./td[6]/text()").getall()
            single['content_6'] = "".join(content_6).strip()
            content_7 = content.xpath("./td[7]/text()").getall()
            single['content_7'] = "".join(content_7).strip()
            content_8 = content.xpath("./td[8]/text()").getall()
            single['content_8'] = "".join(content_8).strip()
            content_9 = content.xpath("./td[9]/text()").getall()
            single['content_9'] = "".join(content_9).strip()
            content_10 = content.xpath("./td[10]/div/text()").getall()
            single['content_10'] = "".join(content_10).strip()
            finance_index1_1.append(single)
            jqka_fi1['finance_index1_1'] = finance_index1_1

        time.sleep(random.randint(2, 5))
        yield jqka_fi1



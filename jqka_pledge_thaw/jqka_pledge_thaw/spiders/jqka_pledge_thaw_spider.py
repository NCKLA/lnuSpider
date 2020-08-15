# -*- coding: utf-8 -*-
import random
import re
# from jqka_com_detail.jqka_com_detail.items import JqkaComDetailItem
from jqka_pledge_thaw.items import JqkaPledgeThawItem
from scrapy import item
from scrapy.http import Request
#获取excel需要引入的包
from openpyxl import load_workbook
from selenium import webdriver
import time
import scrapy


class JqkaPledgeThawSpiderSpider(scrapy.Spider):
    name = 'jqka_pledge_thaw_spider'
    allowed_domains = ['basic.10jqka.com']
    start_urls = ['http://basic.10jqka.com.cn/000002/capital.html']

    def parse(self, response):
        book = load_workbook(filename=r"C:\python\lnuSpider\data\exel\com_list.xlsx")
        sheet = book.active
        data = []
        data1 = []
        data2 = []
        data3 = []
        data4 = []
        row_num = 1
        while row_num <= 3815:
            # 将表中第一列的1-100行数据写入data数组中
            data.append(sheet.cell(row=row_num, column=3).value)
            data1.append(sheet.cell(row=row_num, column=1).value)
            data3.append(sheet.cell(row=row_num, column=2).value)
            data4.append(sheet.cell(row=row_num, column=4).value)
            data2.append(row_num)
            row_num = row_num + 1
        for i in data2:
            # url = 'http://basic.10jqka.com.cn/'+data[i]+'/company.html'
            # print(data[i-1])
            a = str(data[i - 1])
            listedCompany_url = 'http://basic.10jqka.com.cn/' + a + '/capital.html'
            company_pt = JqkaPledgeThawItem()
            company_pt['listedCompany_url'] = listedCompany_url
            listedCompany_id = data1[i - 1]
            company_pt['listedCompany_id'] = listedCompany_id
            listedCompany_name = data3[i - 1]
            company_pt['listedCompany_name'] = listedCompany_name
            listedCompany_fullName = data4[i - 1]
            company_pt['listedCompany_fullName'] = listedCompany_fullName
            yield scrapy.Request(company_pt['listedCompany_url'], meta={'company_pt': company_pt},callback=self.detail_ni, dont_filter=True)

    def detail_ni(self, response):
        company_pt = response.meta['company_pt']

        # 质押解冻
        # listedCompany_capitalOperation_pledgeThaw
        # 首先判断有无改模快
        if response.xpath("//div[@id='freeze']"):
            company_pt['listedCompany_capitalOperation_pledgeThaw'] = list()
            table_list = response.xpath("//div[@id='freeze']//table[@class='m_table']")
            for table in table_list:
                table_dict = dict()
                table_dict['listedCompany_capitalOperation_pledgeThaw_date'] = table.xpath(".//tr[1]/td[1]/em/text()").get()
                table_dict['listedCompany_capitalOperation_pledgeThaw_numberofSharesOriginallyPledged'] = table.xpath(".//tr[1]/td[2]/text()").get()
                table_dict['listedCompany_capitalOperation_pledgeThaw_expectedPledgePeriod'] = table.xpath(".//tr[1]/td[3]/text()").get()
                if table.xpath(".//tr[2]/td/text()[2]"):
                    table_dict['listedCompany_capitalOperation_pledgeThaw_pledgor'] = table.xpath(".//tr[2]/td/text()[2]").get().strip()
                else:
                    table_dict['listedCompany_capitalOperation_pledgeThaw_pledgor'] = table.xpath(".//tr[2]/td/text()[2]").get()
                if table.xpath(".//tr[3]/td/text()[2]"):
                    table_dict['listedCompany_capitalOperation_pledgeThaw_pledgee'] = table.xpath(".//tr[3]/td/text()[2]").get().strip()
                else:
                    table_dict['listedCompany_capitalOperation_pledgeThaw_pledgee'] = table.xpath(".//tr[3]/td/text()[2]").get()
                table_dict['listedCompany_capitalOperation_pledgeThaw_descriptionofPledge'] = table.xpath(".//tr[4]/td/p/text()").get()
                # 加入列表
                company_pt['listedCompany_capitalOperation_pledgeThaw'].append(table_dict)
        else:
            company_pt['listedCompany_capitalOperation_pledgeThaw'] = []


        yield company_pt


import random
import re
from scrapy import item
from scrapy.http import Request
#获取excel需要引入的包
from openpyxl import load_workbook
from selenium import webdriver
import time
import scrapy
from holdings.items import HoldingsItem


class HoldingSpider(scrapy.Spider):
    name = 'holding_spider'
    allowed_domains = ['basic.10.jqka.com']


    # def __init__(self):
    #     #     self.driver = webdriver.PhantomJS(executable_path=r'C:\用户\10359\local\bin\phantomjs.exe')
    #     #     self.driver.set_page_load_timeout(40)

    def start_requests(self):
        yield Request("http://basic.10jqka.com.cn/603221/company.html", headers={
            'User-Agent': "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.87 Safari/537.36"})

    def parse(self, response):
        # 读取excel表格
        book = load_workbook(filename="com_list(2).xlsx")
        sheet = book.active
        data = []
        row_num = 1
        while row_num <= 3:
            # 将表中第一列的1-100行数据写入data数组中
            data.append(sheet.cell(row=row_num, column=3).value)
            row_num = row_num + 1
        for i in data:
            company_share = HoldingsItem()
            url = 'http://basic.10jqka.com.cn/%s' % i + '/company.html'
            company_share['url'] = url
            # print(response.text)
            yield scrapy.Request(company_share['url'],
                                 meta={'company_share': company_share}, callback=self.share, dont_filter=True)
        return

    def share(self, response):
        print("=====准备参股控股公司中的信息====")
        company_share = response.meta['company_share']
        root = response.xpath("//div[@class='content page_event_content']")[0]
        company_name = root.xpath("//div[@stat][@id='detail']//table[@class='m_table']//tr[1]""//td[2]//spa"
                                  "n/text()")[0].extract()
        company_share['company_name'] = company_name

        print("=====信息准备完毕====")

        contents = response.xpath("//div[@id='share'][@class='m_box gssj_scroll'][@stat='company_share']//div"
                                  "[@class='bd pr']/table[@id='ckg_table'][@class='m_table m_hl ggintro business']"
                                  "/tbody/tr")
        company_content1_1 = list()
        for content in contents:
            single = dict()
            content_1 = content.xpath("./td[2]/p/text()").getall()
            single['content_1'] = "".join(content_1).strip()
            content_2 = content.xpath("./td[3]/text()").getall()
            single['content_2'] = "".join(content_2).strip()
            content_3 = content.xpath("./td[4]/text()").getall()
            single['content_3'] = "".join(content_3).strip()
            content_4 = content.xpath("./td[5]/text()").getall()
            single['content_4'] = "".join(content_4).strip()
            content_5 = content.xpath("./td[6]/text()").getall()
            single['content_5'] = "".join(content_5).strip()
            content_6 = content.xpath("./td[7]/text()").getall()
            single['content_6'] = "".join(content_6).strip()
            content_7 = content.xpath("./td[8]/text()").getall()
            single['content_7'] = "".join(content_7).strip()
            company_content1_1.append(single)
            company_share['company_content1_1'] = company_content1_1

        time.sleep(random.randint(2, 5))
        yield company_share








